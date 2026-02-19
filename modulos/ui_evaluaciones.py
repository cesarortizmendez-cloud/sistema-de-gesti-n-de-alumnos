# ============================================  # Separador visual
# modulos/ui_evaluaciones.py                     # Página "Evaluaciones" (UI única, responsiva)
# - Selección de curso (arriba)                  # Combobox de cursos
# - Listado de evaluaciones (izquierda)          # Treeview
# - Formulario CRUD (derecha)                    # Entry + botones
# - Suma de porcentajes por curso                # Validación (verde 100 / rojo diferente)
# - Buscar evaluaciones                          # Filtra por nombre
# - Exportar Excel / PDF                         # Reportes por curso
# - Importar Excel                               # Carga masiva de evaluaciones para el curso seleccionado
# ============================================  # Separador visual

import tkinter as tk                                              # Tkinter base                                                    # noqa: E501
from tkinter import ttk, messagebox, filedialog                   # Widgets + diálogos                                              # noqa: E501

from openpyxl import Workbook                                     # Exportación Excel                                                # noqa: E501
from openpyxl import load_workbook                                # Importación Excel                                                # noqa: E501

from reportlab.pdfgen import canvas                               # PDF                                                              # noqa: E501
from reportlab.lib.pagesizes import A4                             # Tamaño A4                                                        # noqa: E501

from .bd_sqlite import obtener_conexion                            # Conexión SQLite                                                  # noqa: E501
from .config import carpeta_exports                                # Carpeta exportaciones                                            # noqa: E501


class PaginaEvaluaciones(ttk.Frame):                               # Página Evaluaciones                                              # noqa: E501
    def __init__(self, master):                                    # Constructor                                                      # noqa: E501
        super().__init__(master)                                   # Inicializa Frame                                                 # noqa: E501

        self.eval_sel_id = None                                    # ID evaluación seleccionada                                       # noqa: E501
        self.curso_sel_id = None                                   # ID curso seleccionado                                            # noqa: E501

        self._cursos = []                                          # Lista de cursos (dicts)                                          # noqa: E501
        self._cursos_display = []                                  # Lista de strings para combobox                                   # noqa: E501

        self._crear_ui()                                           # Construye la UI                                                  # noqa: E501
        self._cargar_cursos()                                      # Carga cursos al iniciar                                          # noqa: E501

    def on_show(self):                                             # Se llama al mostrar la página (si la principal lo usa)           # noqa: E501
        self._cargar_cursos()                                      # Refresca cursos                                                  # noqa: E501

    # =========================================================
    # Statusbar (si ventana principal tiene set_status)
    # =========================================================
    def _status(self, texto: str):                                 # Envía texto a barra de estado                                    # noqa: E501
        top = self.winfo_toplevel()                                # Ventana principal                                                # noqa: E501
        if hasattr(top, "set_status") and callable(getattr(top, "set_status")):  # Si existe set_status                  # noqa: E501
            top.set_status(texto)                                  # Actualiza estado                                                 # noqa: E501

    # =========================================================
    # Helper SQL
    # =========================================================
    def _q(self, sql: str, params: tuple = (), fetch: bool = False):  # Ejecuta SQL con parámetros                                      # noqa: E501
        conn = obtener_conexion()                                  # Abre conexión SQLite                                             # noqa: E501
        try:                                                       # Bloque protegido                                                 # noqa: E501
            cur = conn.cursor()                                    # Cursor                                                           # noqa: E501
            cur.execute(sql, params)                               # Ejecuta consulta                                                 # noqa: E501
            if fetch:                                              # Si requiere resultados                                           # noqa: E501
                return cur.fetchall()                              # Devuelve filas                                                   # noqa: E501
            conn.commit()                                          # Confirma cambios                                                 # noqa: E501
            return None                                            # Devuelve None                                                    # noqa: E501
        finally:                                                   # Siempre                                                          # noqa: E501
            conn.close()                                           # Cierra conexión                                                  # noqa: E501

    # =========================================================
    # UI
    # =========================================================
    def _crear_ui(self):                                           # Construye interfaz                                               # noqa: E501
        self.columnconfigure(0, weight=1)                          # Frame crece en ancho                                             # noqa: E501
        self.rowconfigure(2, weight=1)                             # ✅ Solo el contenido principal crece en alto                      # noqa: E501

        ttk.Label(self, text="Evaluaciones", font=("Segoe UI", 16, "bold")).grid(  # Título                                 # noqa: E501
            row=0, column=0, sticky="w", padx=12, pady=(12, 6)     # Ubicación                                                        # noqa: E501
        )

        # ---------------- Barra superior (compacta) ----------------
        barra = ttk.Frame(self)                                    # Frame superior                                                   # noqa: E501
        barra.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))  # No se estira en alto                                           # noqa: E501
        barra.columnconfigure(1, weight=1)                         # Combo curso se expande en ancho                                  # noqa: E501
        barra.columnconfigure(5, weight=1)                         # Entry buscar se expande en ancho                                 # noqa: E501

        # Curso
        ttk.Label(barra, text="Curso:").grid(row=0, column=0, sticky="w", padx=(0, 8))  # Label curso                         # noqa: E501
        self.var_curso = tk.StringVar()                            # Variable curso                                                   # noqa: E501
        self.cmb_curso = ttk.Combobox(barra, textvariable=self.var_curso, state="readonly")  # Combo curso                    # noqa: E501
        self.cmb_curso.grid(row=0, column=1, sticky="ew", padx=(0, 10))  # Ubicación                                                     # noqa: E501
        self.cmb_curso.bind("<<ComboboxSelected>>", lambda e: self.on_curso_change())  # Evento cambio                        # noqa: E501

        # Suma porcentajes
        self.lbl_suma = ttk.Label(barra, text="Suma %: 0.00")       # Label suma                                                       # noqa: E501
        self.lbl_suma.grid(row=0, column=2, sticky="w", padx=(0, 12))  # Ubicación                                                     # noqa: E501

        # Buscar
        ttk.Label(barra, text="Buscar:").grid(row=0, column=3, sticky="w", padx=(0, 8))  # Label buscar                       # noqa: E501
        self.var_buscar = tk.StringVar()                            # Variable buscar                                                  # noqa: E501
        ent_buscar = ttk.Entry(barra, textvariable=self.var_buscar) # Entry buscar                                                     # noqa: E501
        ent_buscar.grid(row=0, column=5, sticky="ew", padx=(0, 10))  # Ubicación                                                     # noqa: E501
        ent_buscar.bind("<Return>", lambda e: self.on_buscar())      # Enter aplica                                                     # noqa: E501

        ttk.Button(barra, text="Aplicar", command=self.on_buscar).grid(row=0, column=6, padx=4)   # Botón aplicar             # noqa: E501
        ttk.Button(barra, text="Ver todo", command=self.on_ver_todo).grid(row=0, column=7, padx=4)  # Botón ver todo           # noqa: E501

        # ---------------- Contenido principal: PanedWindow ----------------
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)     # Divisor arrastrable                                              # noqa: E501
        self.paned.grid(row=2, column=0, sticky="nsew", padx=12, pady=12)  # Ocupa todo el contenido                                      # noqa: E501

        # Panel izquierdo (listado)
        self.panel_izq = ttk.Frame(self.paned)                      # Panel izq                                                        # noqa: E501
        self.panel_izq.columnconfigure(0, weight=1)                 # Tabla crece                                                      # noqa: E501
        self.panel_izq.rowconfigure(0, weight=1)                    # Tabla crece                                                      # noqa: E501

        # Panel derecho (formulario)
        self.panel_der = ttk.Frame(self.paned)                      # Panel der                                                        # noqa: E501
        self.panel_der.columnconfigure(0, weight=1)                 # Formulario crece                                                 # noqa: E501
        self.panel_der.rowconfigure(0, weight=1)                    # Formulario crece                                                 # noqa: E501

        self.paned.add(self.panel_izq, weight=3)                    # Proporción izquierda                                             # noqa: E501
        self.paned.add(self.panel_der, weight=2)                    # Proporción derecha                                               # noqa: E501

        try:                                                        # Config mínimos                                                   # noqa: E501
            self.paned.paneconfigure(self.panel_izq, minsize=560)    # Min izq                                                          # noqa: E501
            self.paned.paneconfigure(self.panel_der, minsize=420)    # Min der                                                          # noqa: E501
        except Exception:                                           # Si ttk no soporta                                                # noqa: E501
            pass                                                    # No rompe                                                         # noqa: E501

        # ---------------- IZQUIERDA: Treeview evaluaciones ----------------
        self.tree = ttk.Treeview(                                   # Tabla evaluaciones                                               # noqa: E501
            self.panel_izq,                                         # Padre                                                            # noqa: E501
            columns=("id", "evaluacion", "porcentaje"),             # Columnas                                                         # noqa: E501
            show="headings",                                        # Encabezados                                                      # noqa: E501
            selectmode="browse",                                    # Selección única                                                  # noqa: E501
        )

        self.tree.heading("id", text="ID")                          # Encabezado ID                                                    # noqa: E501
        self.tree.heading("evaluacion", text="Evaluación")          # Encabezado Evaluación                                            # noqa: E501
        self.tree.heading("porcentaje", text="%")                   # Encabezado %                                                     # noqa: E501

        self.tree.column("id", width=70, anchor="center")           # Col ID                                                           # noqa: E501
        self.tree.column("evaluacion", width=340, anchor="w")       # Col Evaluación                                                    # noqa: E501
        self.tree.column("porcentaje", width=90, anchor="center")   # Col %                                                            # noqa: E501

        sb = ttk.Scrollbar(self.panel_izq, orient="vertical", command=self.tree.yview)  # Scroll                         # noqa: E501
        self.tree.configure(yscrollcommand=sb.set)                  # Enlaza scroll                                                     # noqa: E501

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=0)  # Ubicación                                                     # noqa: E501
        sb.grid(row=0, column=1, sticky="ns")                       # Ubicación                                                         # noqa: E501

        self.tree.bind("<<TreeviewSelect>>", lambda e: self.on_select())  # Selección -> formulario                                        # noqa: E501

        # ---------------- DERECHA: Formulario ----------------
        box = ttk.LabelFrame(self.panel_der, text="Formulario")     # Caja formulario                                                  # noqa: E501
        box.grid(row=0, column=0, sticky="nsew")                    # Expandible                                                       # noqa: E501
        box.columnconfigure(1, weight=1)                            # Entradas crecen                                                  # noqa: E501

        # Variables formulario
        self.var_eval_nombre = tk.StringVar()                       # Nombre evaluación                                                # noqa: E501
        self.var_eval_porcentaje = tk.StringVar()                   # Porcentaje                                                       # noqa: E501

        # Curso (solo lectura)
        ttk.Label(box, text="Curso (seleccionado arriba):").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 6))  # Label  # noqa: E501
        self.ent_curso_info = ttk.Entry(box, state="readonly")      # Entry readonly                                                   # noqa: E501
        self.ent_curso_info.grid(row=0, column=1, sticky="ew", padx=10, pady=(10, 6))                                  # Ubicación # noqa: E501

        # Nombre evaluación
        ttk.Label(box, text="Nombre evaluación:").grid(row=1, column=0, sticky="w", padx=10, pady=6)                   # Label     # noqa: E501
        ttk.Entry(box, textvariable=self.var_eval_nombre).grid(row=1, column=1, sticky="ew", padx=10, pady=6)          # Entry     # noqa: E501

        # Porcentaje
        ttk.Label(box, text="Porcentaje (%):").grid(row=2, column=0, sticky="w", padx=10, pady=6)                      # Label     # noqa: E501
        ttk.Entry(box, textvariable=self.var_eval_porcentaje).grid(row=2, column=1, sticky="ew", padx=10, pady=6)      # Entry     # noqa: E501

        # Botones CRUD
        btns = ttk.Frame(box)                                         # Contenedor botones                                               # noqa: E501
        btns.grid(row=3, column=0, columnspan=2, sticky="ew", padx=10, pady=(12, 6))                                   # Ubicación # noqa: E501
        ttk.Button(btns, text="Nuevo", command=self.on_nuevo).pack(side="left", padx=6)                                 # Nuevo    # noqa: E501
        ttk.Button(btns, text="Guardar", command=self.on_guardar).pack(side="left", padx=6)                             # Guardar  # noqa: E501
        ttk.Button(btns, text="Eliminar", command=self.on_eliminar).pack(side="left", padx=6)                           # Eliminar # noqa: E501
        ttk.Button(btns, text="Limpiar", command=self.on_limpiar).pack(side="left", padx=6)                             # Limpiar  # noqa: E501

        ttk.Separator(box, orient="horizontal").grid(row=4, column=0, columnspan=2, sticky="ew", padx=10, pady=10)      # Separador # noqa: E501

        # Botones export/import
        btns2 = ttk.Frame(box)                                         # Frame export/import                                              # noqa: E501
        btns2.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 10))                                   # Ubicación # noqa: E501
        ttk.Button(btns2, text="Exportar Excel", command=self.on_exportar_excel).pack(side="left", padx=6)               # Excel     # noqa: E501
        ttk.Button(btns2, text="Exportar PDF", command=self.on_exportar_pdf).pack(side="left", padx=6)                   # PDF       # noqa: E501
        ttk.Button(btns2, text="Importar Excel", command=self.on_importar_excel).pack(side="left", padx=6)               # Import    # noqa: E501

        ttk.Label(                                                     # Tip                                                               # noqa: E501
            box,
            text="Tip: importe Excel con columnas: Evaluacion | Porcentaje (o %). (Debe haber un curso seleccionado).",
            foreground="gray",
            wraplength=360,
        ).grid(row=6, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 10))                                         # Ubicación # noqa: E501

    # =========================================================
    # Cursos (combobox)
    # =========================================================
    def _cargar_cursos(self):                                         # Carga lista de cursos para el combobox                            # noqa: E501
        rows = self._q(                                               # Query cursos detallados                                           # noqa: E501
            """
            SELECT
              c.curso_id,
              c.nombre AS curso_nombre,
              COALESCE(c.codigo,'') AS codigo,
              c.periodo,
              ca.nombre AS carrera_nombre,
              u.nombre AS universidad_nombre
            FROM cursos c
            JOIN carreras ca ON ca.carrera_id=c.carrera_id
            JOIN universidades u ON u.universidad_id=ca.universidad_id
            ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
            """,
            fetch=True,
        ) or []

        self._cursos = []                                             # Resetea lista                                                     # noqa: E501
        self._cursos_display = []                                     # Resetea display                                                   # noqa: E501

        for r in rows:                                                # Recorre cursos                                                    # noqa: E501
            item = {                                                  # Arma dict                                                        # noqa: E501
                "curso_id": int(r["curso_id"]),                       # ID curso                                                          # noqa: E501
                "universidad_nombre": r["universidad_nombre"],        # Universidad                                                       # noqa: E501
                "carrera_nombre": r["carrera_nombre"],               # Carrera                                                           # noqa: E501
                "periodo": r["periodo"],                              # Periodo                                                           # noqa: E501
                "curso_nombre": r["curso_nombre"],                   # Curso                                                             # noqa: E501
                "codigo": r["codigo"],                                # Código                                                            # noqa: E501
            }
            self._cursos.append(item)                                 # Guarda                                                            # noqa: E501

            txt = f"{item['universidad_nombre']} | {item['carrera_nombre']} | {item['periodo']} | {item['curso_nombre']}"  # Display # noqa: E501
            if item["codigo"]:                                        # Si tiene código                                                   # noqa: E501
                txt += f" ({item['codigo']})"                         # Agrega código                                                     # noqa: E501
            self._cursos_display.append(txt)                          # Guarda display                                                    # noqa: E501

        self.cmb_curso["values"] = self._cursos_display               # Set combobox                                                      # noqa: E501

        if self._cursos_display:                                      # Si hay cursos                                                     # noqa: E501
            if self.var_curso.get().strip() not in self._cursos_display:  # Si selección inválida                                          # noqa: E501
                self.var_curso.set(self._cursos_display[0])           # Selecciona primero                                               # noqa: E501
            self.on_curso_change()                                    # Dispara carga                                                     # noqa: E501
        else:                                                         # Si no hay cursos                                                  # noqa: E501
            self.var_curso.set("")                                    # Limpia                                                            # noqa: E501
            self.curso_sel_id = None                                  # Sin curso                                                         # noqa: E501
            self._actualizar_info_curso("")                           # Limpia entry readonly                                             # noqa: E501
            self._cargar_listado("")                                  # Limpia listado                                                    # noqa: E501
            self._actualizar_suma()                                   # Suma 0                                                            # noqa: E501

    def _curso_info_sel(self):                                        # Retorna dict del curso seleccionado                               # noqa: E501
        txt = self.var_curso.get().strip()                            # Texto actual                                                       # noqa: E501
        if not txt:                                                   # Vacío                                                              # noqa: E501
            return None                                               # None                                                              # noqa: E501
        if txt not in self._cursos_display:                           # Si no existe                                                      # noqa: E501
            return None                                               # None                                                              # noqa: E501
        idx = self._cursos_display.index(txt)                         # Índice                                                             # noqa: E501
        return self._cursos[idx]                                      # Retorna dict                                                       # noqa: E501

    def _actualizar_info_curso(self, texto: str):                     # Escribe texto en Entry readonly                                   # noqa: E501
        self.ent_curso_info.config(state="normal")                    # Habilita temporalmente                                            # noqa: E501
        self.ent_curso_info.delete(0, "end")                          # Limpia                                                            # noqa: E501
        self.ent_curso_info.insert(0, texto)                          # Inserta                                                           # noqa: E501
        self.ent_curso_info.config(state="readonly")                  # Vuelve readonly                                                   # noqa: E501

    def on_curso_change(self):                                        # Evento cambio de curso                                             # noqa: E501
        info = self._curso_info_sel()                                 # Obtiene curso seleccionado                                        # noqa: E501
        if not info:                                                  # Si no hay                                                         # noqa: E501
            self.curso_sel_id = None                                  # Limpia                                                            # noqa: E501
            self._actualizar_info_curso("")                           # Limpia                                                            # noqa: E501
            self._cargar_listado("")                                  # Limpia listado                                                    # noqa: E501
            self._actualizar_suma()                                   # Suma 0                                                            # noqa: E501
            return                                                    # Sale                                                              # noqa: E501

        self.curso_sel_id = int(info["curso_id"])                     # Set curso_id                                                      # noqa: E501
        self.eval_sel_id = None                                       # Resetea selección evaluación                                      # noqa: E501
        self.on_limpiar()                                             # Limpia formulario                                                 # noqa: E501

        texto = f"{info['universidad_nombre']} | {info['carrera_nombre']} | {info['periodo']} | {info['curso_nombre']}"  # Info # noqa: E501
        if info.get("codigo"):                                        # Si tiene código                                                   # noqa: E501
            texto += f" ({info['codigo']})"                           # Agrega                                                            # noqa: E501
        self._actualizar_info_curso(texto)                            # Muestra en readonly                                               # noqa: E501

        self._cargar_listado(self.var_buscar.get().strip())           # Carga evaluaciones                                                # noqa: E501
        self._actualizar_suma()                                       # Actualiza suma                                                    # noqa: E501

    # =========================================================
    # Listado evaluaciones
    # =========================================================
    def _actualizar_suma(self):                                       # Actualiza label suma %                                             # noqa: E501
        if self.curso_sel_id is None:                                 # Sin curso                                                         # noqa: E501
            self.lbl_suma.config(text="Suma %: 0.00", foreground="black")  # Texto                                                         # noqa: E501
            return                                                    # Sale                                                              # noqa: E501

        row = self._q(                                                # SUM porcentajes                                                   # noqa: E501
            "SELECT COALESCE(SUM(porcentaje),0) AS s FROM evaluaciones WHERE curso_id=?",
            (self.curso_sel_id,),
            fetch=True,
        )
        s = float(row[0]["s"] if row else 0)                          # Suma                                                              # noqa: E501
        ok = abs(s - 100.0) < 0.001                                   # True si suma=100                                                  # noqa: E501
        self.lbl_suma.config(text=f"Suma %: {s:.2f}", foreground=("green" if ok else "red"))  # Color                                        # noqa: E501

    def _cargar_listado(self, texto: str):                             # Carga tabla evaluaciones                                          # noqa: E501
        for i in self.tree.get_children():                             # Limpia                                                            # noqa: E501
            self.tree.delete(i)                                        # Borra                                                             # noqa: E501

        if self.curso_sel_id is None:                                  # Sin curso seleccionado                                            # noqa: E501
            self._status("Vista: Evaluaciones | Seleccione un curso")  # Status                                                            # noqa: E501
            return                                                     # Sale                                                              # noqa: E501

        t = (texto or "").strip()                                      # Texto búsqueda                                                     # noqa: E501

        if t:                                                          # Con filtro                                                         # noqa: E501
            q = f"%{t}%"                                              # Like                                                              # noqa: E501
            rows = self._q(
                """
                SELECT evaluacion_id AS id, nombre, porcentaje
                FROM evaluaciones
                WHERE curso_id=? AND nombre LIKE ? COLLATE NOCASE
                ORDER BY evaluacion_id
                """,
                (self.curso_sel_id, q),
                fetch=True,
            ) or []
        else:                                                          # Sin filtro                                                         # noqa: E501
            rows = self._q(
                """
                SELECT evaluacion_id AS id, nombre, porcentaje
                FROM evaluaciones
                WHERE curso_id=?
                ORDER BY evaluacion_id
                """,
                (self.curso_sel_id,),
                fetch=True,
            ) or []

        for r in rows:                                                  # Inserta filas                                                      # noqa: E501
            self.tree.insert(                                           # Inserta                                                           # noqa: E501
                "",
                "end",
                values=(int(r["id"]), r["nombre"], f"{float(r['porcentaje']):.2f}"),
            )

        self._status(f"Vista: Evaluaciones | CursoID: {self.curso_sel_id} | Registros: {len(rows)}")  # Status                   # noqa: E501

    def on_buscar(self):                                                # Aplica filtro                                                      # noqa: E501
        self._cargar_listado(self.var_buscar.get().strip())             # Recarga                                                            # noqa: E501

    def on_ver_todo(self):                                              # Limpia filtro                                                      # noqa: E501
        self.var_buscar.set("")                                         # Limpia                                                             # noqa: E501
        self._cargar_listado("")                                        # Carga sin filtro                                                   # noqa: E501

    # =========================================================
    # Selección -> formulario
    # =========================================================
    def on_select(self):                                                # Carga selección al formulario                                      # noqa: E501
        sel = self.tree.selection()                                     # Selección                                                          # noqa: E501
        if not sel:                                                     # Si no hay                                                          # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        vals = self.tree.item(sel[0], "values")                         # Valores                                                            # noqa: E501
        if not vals:                                                    # Si vacío                                                           # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        self.eval_sel_id = int(vals[0])                                 # evaluacion_id                                                      # noqa: E501
        self.var_eval_nombre.set(vals[1])                               # nombre evaluación                                                  # noqa: E501
        self.var_eval_porcentaje.set(str(vals[2]))                      # porcentaje                                                         # noqa: E501

    # =========================================================
    # CRUD
    # =========================================================
    def on_nuevo(self):                                                 # Nuevo (limpia selección)                                           # noqa: E501
        self.eval_sel_id = None                                         # Sin eval seleccionada                                              # noqa: E501
        self.on_limpiar()                                               # Limpia campos                                                      # noqa: E501

    def on_limpiar(self):                                               # Limpia formulario                                                  # noqa: E501
        self.var_eval_nombre.set("")                                    # Limpia nombre                                                      # noqa: E501
        self.var_eval_porcentaje.set("")                                # Limpia porcentaje                                                  # noqa: E501

    def on_guardar(self):                                               # Guarda (INSERT/UPDATE)                                             # noqa: E501
        if self.curso_sel_id is None:                                   # Debe haber curso                                                   # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un curso.")   # Aviso                                                              # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        nombre = self.var_eval_nombre.get().strip()                     # Nombre                                                             # noqa: E501
        por_str = self.var_eval_porcentaje.get().strip()                # Porcentaje string                                                  # noqa: E501

        if not nombre:                                                  # Validación                                                         # noqa: E501
            messagebox.showwarning("Atención", "El nombre de la evaluación es obligatorio.")  # Aviso                     # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        try:
            por = float(por_str.replace(",", "."))                      # Convierte (soporta coma)                                           # noqa: E501
        except Exception:
            messagebox.showwarning("Atención", "Porcentaje inválido (debe ser numérico).")    # Aviso                     # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        if por <= 0 or por > 100:                                       # Rango válido                                                      # noqa: E501
            messagebox.showwarning("Atención", "Porcentaje debe estar entre 0 y 100.")        # Aviso                     # noqa: E501
            return                                                      # Sale                                                               # noqa: E501

        try:
            if self.eval_sel_id is None:                                # INSERT                                                            # noqa: E501
                self._q(
                    "INSERT INTO evaluaciones(curso_id, nombre, porcentaje) VALUES(?,?,?)",
                    (self.curso_sel_id, nombre, por),
                )
                messagebox.showinfo("OK", "Evaluación creada.")         # OK                                                                # noqa: E501
            else:                                                       # UPDATE                                                            # noqa: E501
                self._q(
                    "UPDATE evaluaciones SET nombre=?, porcentaje=? WHERE evaluacion_id=?",
                    (nombre, por, self.eval_sel_id),
                )
                messagebox.showinfo("OK", "Evaluación actualizada.")    # OK                                                                # noqa: E501

            self._cargar_listado(self.var_buscar.get().strip())         # Refresca tabla                                                     # noqa: E501
            self._actualizar_suma()                                     # Actualiza suma                                                     # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                       # Error                                                             # noqa: E501

    def on_eliminar(self):                                              # Elimina evaluación                                                 # noqa: E501
        if self.eval_sel_id is None:                                    # Debe seleccionar                                                   # noqa: E501
            messagebox.showwarning("Atención", "Seleccione una evaluación para eliminar.")  # Aviso                     # noqa: E501
            return                                                      # Sale                                                               # noqa: E501
        if not messagebox.askyesno("Confirmar", "¿Eliminar evaluación? (las notas asociadas se eliminarán por cascada)"):  # Confirm               # noqa: E501
            return                                                      # Cancel                                                             # noqa: E501

        try:
            self._q("DELETE FROM evaluaciones WHERE evaluacion_id=?", (self.eval_sel_id,))  # Delete                    # noqa: E501
            self.eval_sel_id = None                                    # Limpia selección                                                   # noqa: E501
            self.on_limpiar()                                          # Limpia formulario                                                  # noqa: E501
            self._cargar_listado(self.var_buscar.get().strip())        # Refresca tabla                                                     # noqa: E501
            self._actualizar_suma()                                    # Refresca suma                                                      # noqa: E501
            messagebox.showinfo("OK", "Evaluación eliminada.")         # OK                                                                 # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                      # Error                                                              # noqa: E501

    # =========================================================
    # Exportaciones (por curso)
    # =========================================================
    def _datos_reporte(self):                                          # Obtiene datos para exportar                                        # noqa: E501
        if self.curso_sel_id is None:                                  # Debe haber curso                                                   # noqa: E501
            raise ValueError("Debe seleccionar un curso.")             # Error                                                              # noqa: E501

        # Curso info (para encabezado)
        info = self._curso_info_sel()                                  # Dict curso                                                         # noqa: E501
        if not info:                                                   # Si no hay                                                          # noqa: E501
            raise ValueError("No se pudo obtener información del curso.")  # Error                                                          # noqa: E501

        # Evaluaciones del curso (filtradas por búsqueda actual)
        t = self.var_buscar.get().strip()                              # Texto                                                              # noqa: E501
        if t:                                                          # Con filtro                                                         # noqa: E501
            q = f"%{t}%"                                              # Like                                                               # noqa: E501
            rows = self._q(
                "SELECT evaluacion_id, nombre, porcentaje FROM evaluaciones WHERE curso_id=? AND nombre LIKE ? COLLATE NOCASE ORDER BY evaluacion_id",
                (self.curso_sel_id, q),
                fetch=True,
            ) or []
        else:                                                          # Sin filtro                                                         # noqa: E501
            rows = self._q(
                "SELECT evaluacion_id, nombre, porcentaje FROM evaluaciones WHERE curso_id=? ORDER BY evaluacion_id",
                (self.curso_sel_id,),
                fetch=True,
            ) or []

        # Suma porcentajes
        srow = self._q("SELECT COALESCE(SUM(porcentaje),0) AS s FROM evaluaciones WHERE curso_id=?", (self.curso_sel_id,), fetch=True)  # SUM  # noqa: E501
        suma = float(srow[0]["s"] if srow else 0)                      # Suma                                                              # noqa: E501

        return info, rows, suma                                        # Retorna todo                                                      # noqa: E501

    def on_exportar_excel(self):                                       # Exporta evaluaciones a Excel                                       # noqa: E501
        try:
            info, rows, suma = self._datos_reporte()                   # Datos                                                              # noqa: E501

            ruta = filedialog.asksaveasfilename(                       # Guardar                                                            # noqa: E501
                title="Guardar Excel (Evaluaciones)",                  # Título                                                             # noqa: E501
                defaultextension=".xlsx",                              # Extensión                                                          # noqa: E501
                filetypes=[("Excel", "*.xlsx")],                       # Tipos                                                               # noqa: E501
                initialdir=carpeta_exports(),                          # Carpeta sugerida                                                   # noqa: E501
            )
            if not ruta:                                               # Cancel                                                              # noqa: E501
                return                                                 # Sale                                                                # noqa: E501

            wb = Workbook()                                            # Workbook                                                           # noqa: E501
            ws = wb.active                                             # Hoja                                                               # noqa: E501
            ws.title = "Evaluaciones"                                  # Nombre hoja                                                        # noqa: E501

            # Encabezado informativo
            ws.append(["Universidad", info["universidad_nombre"]])     # Universidad                                                        # noqa: E501
            ws.append(["Carrera", info["carrera_nombre"]])             # Carrera                                                            # noqa: E501
            ws.append(["Periodo", info["periodo"]])                    # Periodo                                                            # noqa: E501
            ws.append(["Curso", info["curso_nombre"]])                 # Curso                                                              # noqa: E501
            ws.append(["Codigo", info.get("codigo", "")])              # Código                                                             # noqa: E501
            ws.append(["Suma porcentajes", f"{suma:.2f}"])             # Suma                                                              # noqa: E501
            ws.append([])                                              # Línea en blanco                                                    # noqa: E501

            # Tabla
            ws.append(["EvalID", "Evaluacion", "Porcentaje"])          # Cabecera                                                           # noqa: E501
            for r in rows:                                             # Filas                                                              # noqa: E501
                ws.append([int(r["evaluacion_id"]), r["nombre"], float(r["porcentaje"])])  # Inserta                                          # noqa: E501

            wb.save(ruta)                                              # Guarda                                                             # noqa: E501
            messagebox.showinfo("OK", "Excel exportado correctamente.")  # OK                                                              # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                      # Error                                                              # noqa: E501

    def on_exportar_pdf(self):                                         # Exporta evaluaciones a PDF                                         # noqa: E501
        try:
            info, rows, suma = self._datos_reporte()                   # Datos                                                              # noqa: E501

            ruta = filedialog.asksaveasfilename(                       # Guardar                                                            # noqa: E501
                title="Guardar PDF (Evaluaciones)",                    # Título                                                             # noqa: E501
                defaultextension=".pdf",                               # Extensión                                                          # noqa: E501
                filetypes=[("PDF", "*.pdf")],                          # Tipos                                                              # noqa: E501
                initialdir=carpeta_exports(),                          # Carpeta sugerida                                                   # noqa: E501
            )
            if not ruta:                                               # Cancel                                                              # noqa: E501
                return                                                 # Sale                                                                # noqa: E501

            c = canvas.Canvas(ruta, pagesize=A4)                       # Canvas                                                             # noqa: E501
            w, h = A4                                                  # Dimensiones                                                        # noqa: E501
            y = h - 60                                                 # Cursor Y                                                           # noqa: E501

            c.setFont("Helvetica-Bold", 13)                            # Fuente título                                                      # noqa: E501
            c.drawString(40, y, "Evaluaciones del curso")              # Título                                                             # noqa: E501
            y -= 18                                                    # Baja                                                                # noqa: E501

            c.setFont("Helvetica", 10)                                 # Fuente normal                                                      # noqa: E501
            c.drawString(40, y, f"{info['universidad_nombre']} | {info['carrera_nombre']} | {info['periodo']} | {info['curso_nombre']}")  # Info # noqa: E501
            y -= 14                                                    # Baja                                                                # noqa: E501
            if info.get("codigo"):                                     # Si hay código                                                       # noqa: E501
                c.drawString(40, y, f"Código: {info['codigo']}")       # Código                                                             # noqa: E501
                y -= 14                                                # Baja                                                                # noqa: E501
            c.drawString(40, y, f"Suma porcentajes: {suma:.2f}%")      # Suma                                                               # noqa: E501
            y -= 22                                                    # Baja                                                                # noqa: E501

            c.setFont("Helvetica-Bold", 10)                            # Cabecera                                                           # noqa: E501
            c.drawString(40, y, "EvalID")                              # Col                                                                # noqa: E501
            c.drawString(100, y, "Evaluación")                         # Col                                                                # noqa: E501
            c.drawString(520, y, "%")                                  # Col                                                                # noqa: E501
            y -= 12                                                    # Baja                                                                # noqa: E501

            c.setFont("Helvetica", 10)                                 # Fuente normal                                                      # noqa: E501
            for r in rows:                                             # Filas                                                              # noqa: E501
                if y < 60:                                             # Salto página                                                       # noqa: E501
                    c.showPage()                                      # Nueva                                                              # noqa: E501
                    y = h - 60                                        # Reinicia                                                           # noqa: E501
                    c.setFont("Helvetica", 10)                        # Fuente                                                             # noqa: E501

                c.drawString(40, y, str(int(r["evaluacion_id"])))      # ID                                                                 # noqa: E501
                c.drawString(100, y, str(r["nombre"])[:55])            # Nombre                                                             # noqa: E501
                c.drawRightString(560, y, f"{float(r['porcentaje']):.2f}")  # %                                                             # noqa: E501
                y -= 14                                                # Baja                                                                # noqa: E501

            c.save()                                                   # Guarda                                                             # noqa: E501
            messagebox.showinfo("OK", "PDF exportado correctamente.")   # OK                                                                 # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                      # Error                                                              # noqa: E501

    # =========================================================
    # Importar Excel (Evaluaciones) para curso seleccionado
    # Requiere columnas:
    #   Evaluacion | Porcentaje (o %)   (cabecera en fila 1)
    # =========================================================
    def on_importar_excel(self):                                      # Importa evaluaciones desde Excel                                  # noqa: E501
        if self.curso_sel_id is None:                                  # Debe seleccionar curso                                             # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un curso antes de importar evaluaciones.")  # Aviso                # noqa: E501
            return                                                     # Sale                                                               # noqa: E501

        ruta = filedialog.askopenfilename(                             # Seleccionar archivo                                                # noqa: E501
            title="Importar Excel (Evaluaciones)",                     # Título                                                             # noqa: E501
            filetypes=[("Excel", "*.xlsx")],                           # Tipos                                                               # noqa: E501
            initialdir=".",                                            # Carpeta inicial                                                    # noqa: E501
        )
        if not ruta:                                                   # Cancel                                                              # noqa: E501
            return                                                     # Sale                                                                # noqa: E501

        try:
            wb = load_workbook(ruta)                                   # Abre libro                                                         # noqa: E501
            ws = wb.active                                             # Hoja activa                                                        # noqa: E501

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Cabecera                            # noqa: E501
            headers = [str(h or "").strip().casefold() for h in header_row]          # Normaliza                            # noqa: E501

            def hidx(*names: str) -> int:                              # Encuentra índice de columna                                       # noqa: E501
                for n in names:                                        # Recorre alias                                                     # noqa: E501
                    k = n.casefold()                                   # Normaliza                                                         # noqa: E501
                    if k in headers:                                   # Existe                                                            # noqa: E501
                        return headers.index(k)                        # Retorna índice                                                    # noqa: E501
                return -1                                              # No encontrado                                                     # noqa: E501

            i_nom = hidx("evaluacion", "evaluación", "nombre")         # Col nombre                                                        # noqa: E501
            i_por = hidx("porcentaje", "%", "porc")                    # Col porcentaje                                                    # noqa: E501

            if i_nom == -1 or i_por == -1:                             # Validación                                                        # noqa: E501
                raise ValueError("El Excel debe incluir columnas: Evaluacion y Porcentaje (o %).")  # Error                   # noqa: E501

            conn = obtener_conexion()                                  # Una conexión para velocidad                                        # noqa: E501
            try:
                cur = conn.cursor()                                    # Cursor                                                             # noqa: E501

                total = 0                                              # Total filas                                                        # noqa: E501
                ok = 0                                                 # OK                                                                 # noqa: E501
                errores = []                                           # Errores                                                            # noqa: E501

                for fila_n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Datos                    # noqa: E501
                    total += 1                                         # Cuenta                                                             # noqa: E501
                    try:
                        nombre = str(row[i_nom] or "").strip()         # Nombre                                                             # noqa: E501
                        por_raw = str(row[i_por] or "").strip()        # Porcentaje                                                         # noqa: E501

                        if not nombre:                                 # Obligatorio                                                        # noqa: E501
                            raise ValueError("Nombre de evaluación vacío.")  # Error                                                       # noqa: E501

                        por = float(por_raw.replace(",", "."))         # Convierte                                                          # noqa: E501
                        if por <= 0 or por > 100:                      # Validación rango                                                   # noqa: E501
                            raise ValueError("Porcentaje fuera de rango (0-100].")  # Error                                           # noqa: E501

                        # UPSERT por UNIQUE(curso_id, nombre)
                        cur.execute(                                   # Inserta o actualiza porcentaje                                     # noqa: E501
                            """
                            INSERT INTO evaluaciones(curso_id, nombre, porcentaje)
                            VALUES(?,?,?)
                            ON CONFLICT(curso_id, nombre)
                            DO UPDATE SET porcentaje=excluded.porcentaje
                            """,
                            (self.curso_sel_id, nombre, por),
                        )

                        ok += 1                                        # OK                                                                 # noqa: E501
                    except Exception as e:
                        errores.append(f"Fila {fila_n}: {e}")          # Guarda error                                                       # noqa: E501

                conn.commit()                                          # Confirma                                                           # noqa: E501
            finally:
                conn.close()                                           # Cierra                                                             # noqa: E501

            # Refresca UI
            self._cargar_listado(self.var_buscar.get().strip())         # Recarga tabla                                                      # noqa: E501
            self._actualizar_suma()                                     # Suma                                                               # noqa: E501

            msg = f"Importación finalizada.\nProcesadas: {total}\nOK: {ok}\nErrores: {len(errores)}"  # Resumen               # noqa: E501
            if errores:                                                 # Si hay errores                                                     # noqa: E501
                msg += "\n\nPrimeros errores:\n" + "\n".join(errores[:10])  # Muestra 10                                                    # noqa: E501
            messagebox.showinfo("Resultado", msg)                       # Muestra                                                            # noqa: E501

        except Exception as e:
            messagebox.showerror("Error", str(e))                       # Error general                                                      # noqa: E501
