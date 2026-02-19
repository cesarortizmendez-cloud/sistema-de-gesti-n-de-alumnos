# ============================================  # Separador visual
# modulos/ui_cursos.py                           # Página "Cursos" (UI única, responsiva)
# - Izquierda: listado de cursos                 # Treeview con cursos
# - Derecha: formulario CRUD                     # Universidad, carrera, periodo, nombre, código
# - Barra superior compacta: buscar + export/import  # Acciones rápidas
# - Importación Excel: crea universidades/carreras si no existen  # Carga masiva
# ============================================  # Separador visual

import tkinter as tk                                              # Tkinter base
from tkinter import ttk, messagebox, filedialog, simpledialog      # Widgets + diálogos

from openpyxl import Workbook                                     # Exportar Excel
from openpyxl import load_workbook                                # Importar Excel

from reportlab.pdfgen import canvas                               # Generar PDF
from reportlab.lib.pagesizes import A4                             # Tamaño A4

from .bd_sqlite import obtener_conexion                           # Conexión SQLite
from .config import carpeta_exports                               # Carpeta exports


class PaginaCursos(ttk.Frame):                                    # Página Cursos (Frame embebido)
    def __init__(self, master):                                   # Constructor
        super().__init__(master)                                  # Inicializa Frame

        self.curso_sel_id = None                                  # ID del curso seleccionado (Treeview)
        self._universidades = []                                  # Lista dict universidades
        self._carreras = []                                       # Lista dict carreras (según universidad)
        self._uni_display = []                                    # Strings para combobox universidad
        self._car_display = []                                    # Strings para combobox carrera

        self._crear_ui()                                          # Crea interfaz
        self._cargar_universidades()                              # Carga universidades
        self._cargar_listado("")                                  # Carga listado completo

    def on_show(self):                                            # Hook cuando se muestra la página
        self._cargar_universidades()                              # Refresca universidades (por si cambiaron)
        self._cargar_listado(self.var_buscar.get().strip())        # Refresca listado con filtro actual

    # =========================================================
    # Statusbar (si ventana principal tiene set_status)
    # =========================================================
    def _status(self, texto: str):                                # Envía texto a barra de estado
        top = self.winfo_toplevel()                               # Ventana principal
        if hasattr(top, "set_status") and callable(getattr(top, "set_status")):  # Si existe set_status
            top.set_status(texto)                                 # Actualiza estado

    # =========================================================
    # Helper SQL
    # =========================================================
    def _q(self, sql: str, params: tuple = (), fetch: bool = False):  # Ejecuta SQL con parámetros
        conn = obtener_conexion()                                 # Abre conexión SQLite
        try:                                                      # Bloque seguro
            cur = conn.cursor()                                   # Cursor
            cur.execute(sql, params)                              # Ejecuta
            if fetch:                                             # Si se requieren resultados
                return cur.fetchall()                             # Devuelve filas
            conn.commit()                                         # Confirma cambios
            return None                                           # Devuelve None
        finally:                                                  # Siempre
            conn.close()                                          # Cierra conexión

    # =========================================================
    # UI
    # =========================================================
    def _crear_ui(self):                                          # Construye UI
        self.columnconfigure(0, weight=1)                         # Frame crece en ancho
        self.rowconfigure(2, weight=1)                            # ✅ Solo contenido crece en alto

        ttk.Label(self, text="Cursos", font=("Segoe UI", 16, "bold")).grid(  # Título
            row=0, column=0, sticky="w", padx=12, pady=(12, 6)
        )

        # ---------------- Barra superior compacta ----------------
        barra = ttk.Frame(self)                                   # Frame superior
        barra.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        barra.columnconfigure(1, weight=1)                        # Buscar se expande

        ttk.Label(barra, text="Buscar:").grid(row=0, column=0, sticky="w")  # Label buscar

        self.var_buscar = tk.StringVar()                           # Variable buscar
        ent = ttk.Entry(barra, textvariable=self.var_buscar)       # Entry buscar
        ent.grid(row=0, column=1, sticky="ew", padx=(8, 10))        # Ubicación
        ent.bind("<Return>", lambda e: self.on_buscar())            # Enter aplica búsqueda

        ttk.Button(barra, text="Aplicar", command=self.on_buscar).grid(row=0, column=2, padx=4)  # Botón aplicar
        ttk.Button(barra, text="Ver todo", command=self.on_ver_todo).grid(row=0, column=3, padx=4)  # Botón ver todo

        ttk.Separator(barra, orient="vertical").grid(row=0, column=4, sticky="ns", padx=10)       # Separador

        ttk.Button(barra, text="Exportar Excel", command=self.on_exportar_excel).grid(row=0, column=5, padx=4)  # Excel
        ttk.Button(barra, text="Exportar PDF", command=self.on_exportar_pdf).grid(row=0, column=6, padx=4)      # PDF
        ttk.Button(barra, text="Importar Excel", command=self.on_importar_excel).grid(row=0, column=7, padx=4)  # Import

        # ---------------- Contenido principal: PanedWindow ----------------
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)    # Divisor izq/der
        self.paned.grid(row=2, column=0, sticky="nsew", padx=12, pady=12)

        self.panel_izq = ttk.Frame(self.paned)                     # Panel izquierdo
        self.panel_der = ttk.Frame(self.paned)                     # Panel derecho

        self.panel_izq.columnconfigure(0, weight=1)                # Tabla crece
        self.panel_izq.rowconfigure(0, weight=1)                   # Tabla crece

        self.panel_der.columnconfigure(0, weight=1)                # Form crece
        self.panel_der.rowconfigure(0, weight=1)                   # Form crece

        self.paned.add(self.panel_izq, weight=3)                   # Izq más ancho
        self.paned.add(self.panel_der, weight=2)                   # Der

        try:                                                       # Config mínimos panes
            self.paned.paneconfigure(self.panel_izq, minsize=620)   # Min izq
            self.paned.paneconfigure(self.panel_der, minsize=420)   # Min der
        except Exception:                                          # Si no soporta
            pass                                                   # No rompe

        # ---------------- IZQUIERDA: listado de cursos ----------------
        self.tree = ttk.Treeview(                                  # Treeview cursos
            self.panel_izq,
            columns=("id", "uni", "car", "per", "curso", "cod"),
            show="headings",
            selectmode="browse",
        )

        self.tree.heading("id", text="ID")                         # Encabezado
        self.tree.heading("uni", text="Universidad")               # Encabezado
        self.tree.heading("car", text="Carrera")                   # Encabezado
        self.tree.heading("per", text="Periodo")                   # Encabezado
        self.tree.heading("curso", text="Curso")                   # Encabezado
        self.tree.heading("cod", text="Código")                    # Encabezado

        self.tree.column("id", width=70, anchor="center")          # Col ID
        self.tree.column("uni", width=200, anchor="w")             # Col uni
        self.tree.column("car", width=200, anchor="w")             # Col carrera
        self.tree.column("per", width=90, anchor="center")         # Col periodo
        self.tree.column("curso", width=260, anchor="w")           # Col curso
        self.tree.column("cod", width=110, anchor="center")        # Col código

        sb = ttk.Scrollbar(self.panel_izq, orient="vertical", command=self.tree.yview)  # Scroll
        self.tree.configure(yscrollcommand=sb.set)                 # Enlaza scroll

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 6)) # Ubicación tree
        sb.grid(row=0, column=1, sticky="ns")                      # Ubicación scroll

        self.tree.bind("<<TreeviewSelect>>", lambda e: self.on_select())  # Selección -> formulario

        # ---------------- DERECHA: formulario CRUD ----------------
        box = ttk.LabelFrame(self.panel_der, text="Formulario")    # Marco formulario
        box.grid(row=0, column=0, sticky="nsew")                   # Expande
        box.columnconfigure(1, weight=1)                           # Entradas crecen

        # Variables formulario
        self.var_uni = tk.StringVar()                              # Universidad seleccionada (texto)
        self.var_car = tk.StringVar()                              # Carrera seleccionada (texto)
        self.var_periodo = tk.StringVar()                          # Periodo (YYYY-1 o YYYY-2)
        self.var_nombre = tk.StringVar()                           # Nombre curso
        self.var_codigo = tk.StringVar()                           # Código curso

        # Universidad
        ttk.Label(box, text="Universidad:").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 6))  # Label
        fila_uni = ttk.Frame(box)                                  # Frame para combo+botón
        fila_uni.grid(row=0, column=1, sticky="ew", padx=10, pady=(10, 6))                              # Ubicación
        fila_uni.columnconfigure(0, weight=1)                      # Combo crece

        self.cmb_uni = ttk.Combobox(fila_uni, textvariable=self.var_uni, state="readonly")            # Combo universidad
        self.cmb_uni.grid(row=0, column=0, sticky="ew")            # Ubicación combo
        self.cmb_uni.bind("<<ComboboxSelected>>", lambda e: self.on_uni_change())                      # Evento cambio

        ttk.Button(fila_uni, text="+", width=3, command=self.on_nueva_universidad).grid(row=0, column=1, padx=(6, 0))  # Botón +

        # Carrera
        ttk.Label(box, text="Carrera:").grid(row=1, column=0, sticky="w", padx=10, pady=6)            # Label
        fila_car = ttk.Frame(box)                                  # Frame combo+botón
        fila_car.grid(row=1, column=1, sticky="ew", padx=10, pady=6)                                    # Ubicación
        fila_car.columnconfigure(0, weight=1)                      # Combo crece

        self.cmb_car = ttk.Combobox(fila_car, textvariable=self.var_car, state="readonly")            # Combo carrera
        self.cmb_car.grid(row=0, column=0, sticky="ew")            # Ubicación combo
        ttk.Button(fila_car, text="+", width=3, command=self.on_nueva_carrera).grid(row=0, column=1, padx=(6, 0))     # Botón +

        # Periodo
        ttk.Label(box, text="Periodo (YYYY-1 / YYYY-2):").grid(row=2, column=0, sticky="w", padx=10, pady=6)          # Label
        ttk.Entry(box, textvariable=self.var_periodo).grid(row=2, column=1, sticky="ew", padx=10, pady=6)            # Entry

        # Nombre curso
        ttk.Label(box, text="Nombre del curso:").grid(row=3, column=0, sticky="w", padx=10, pady=6)                  # Label
        ttk.Entry(box, textvariable=self.var_nombre).grid(row=3, column=1, sticky="ew", padx=10, pady=6)            # Entry

        # Código
        ttk.Label(box, text="Código (opcional):").grid(row=4, column=0, sticky="w", padx=10, pady=6)                 # Label
        ttk.Entry(box, textvariable=self.var_codigo).grid(row=4, column=1, sticky="ew", padx=10, pady=6)            # Entry

        # Botones CRUD
        btns = ttk.Frame(box)                                       # Frame botones
        btns.grid(row=5, column=0, columnspan=2, sticky="ew", padx=10, pady=(12, 6))                                  # Ubicación
        ttk.Button(btns, text="Nuevo", command=self.on_nuevo).pack(side="left", padx=6)                               # Nuevo
        ttk.Button(btns, text="Guardar", command=self.on_guardar).pack(side="left", padx=6)                           # Guardar
        ttk.Button(btns, text="Eliminar", command=self.on_eliminar).pack(side="left", padx=6)                         # Eliminar
        ttk.Button(btns, text="Limpiar", command=self.on_limpiar).pack(side="left", padx=6)                           # Limpiar

        ttk.Separator(box, orient="horizontal").grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=10)    # Separador

        ttk.Label(                                                    # Tip de importación
            box,
            text="Importar Excel: columnas sugeridas -> Universidad | Carrera | Periodo | Curso | Codigo",
            foreground="gray",
            wraplength=360,
        ).grid(row=7, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 10))                                      # Ubicación

    # =========================================================
    # Cargar combos (universidades y carreras)
    # =========================================================
    def _cargar_universidades(self):                                 # Carga universidades desde BD
        rows = self._q("SELECT universidad_id, nombre FROM universidades ORDER BY nombre", fetch=True) or []  # Query
        self._universidades = [{"id": int(r["universidad_id"]), "nombre": r["nombre"]} for r in rows]         # Lista dict
        self._uni_display = [u["nombre"] for u in self._universidades]                                         # Display
        self.cmb_uni["values"] = self._uni_display                                                   # Set combo

        if self._uni_display:                                       # Si hay universidades
            if self.var_uni.get().strip() not in self._uni_display: # Si selección inválida
                self.var_uni.set(self._uni_display[0])              # Selecciona primera
            self.on_uni_change()                                    # Carga carreras
        else:                                                       # Si no hay universidades
            self.var_uni.set("")                                    # Limpia
            self.var_car.set("")                                    # Limpia
            self.cmb_car["values"] = []                             # Limpia carreras
            self._carreras = []                                     # Lista vacía

    def _uni_info_sel(self):                                        # Retorna dict universidad seleccionada
        txt = self.var_uni.get().strip()                            # Texto seleccionado
        if not txt or txt not in self._uni_display:                 # Validación
            return None                                             # None
        idx = self._uni_display.index(txt)                          # Índice
        return self._universidades[idx]                             # Dict universidad

    def _cargar_carreras(self, universidad_id: int):                # Carga carreras para una universidad
        rows = self._q(
            "SELECT carrera_id, nombre FROM carreras WHERE universidad_id=? ORDER BY nombre",
            (universidad_id,),
            fetch=True,
        ) or []
        self._carreras = [{"id": int(r["carrera_id"]), "nombre": r["nombre"]} for r in rows]       # Lista dict
        self._car_display = [c["nombre"] for c in self._carreras]                                  # Display
        self.cmb_car["values"] = self._car_display                                                 # Set combo

        if self._car_display:                                      # Si hay carreras
            if self.var_car.get().strip() not in self._car_display:# Si selección inválida
                self.var_car.set(self._car_display[0])             # Selecciona primera
        else:                                                      # Si no hay carreras
            self.var_car.set("")                                   # Limpia

    def _car_info_sel(self):                                       # Retorna dict carrera seleccionada
        txt = self.var_car.get().strip()                           # Texto
        if not txt or txt not in self._car_display:                # Validación
            return None                                            # None
        idx = self._car_display.index(txt)                         # Índice
        return self._carreras[idx]                                 # Dict carrera

    def on_uni_change(self):                                       # Cuando cambia universidad
        uni = self._uni_info_sel()                                 # Dict uni
        if not uni:                                                # Si no hay
            self.cmb_car["values"] = []                            # Limpia combo carreras
            self._carreras = []                                    # Limpia lista
            self.var_car.set("")                                   # Limpia var
            return                                                 # Sale
        self._cargar_carreras(uni["id"])                           # Carga carreras según uni

    # =========================================================
    # Listado / búsqueda
    # =========================================================
    def _cargar_listado(self, texto: str):                          # Carga cursos en la tabla
        for i in self.tree.get_children():                          # Limpia tabla
            self.tree.delete(i)                                     # Borra item

        t = (texto or "").strip()                                   # Texto filtro

        if t:                                                       # Si hay filtro
            q = f"%{t}%"                                           # Like
            rows = self._q(
                """
                SELECT
                  c.curso_id,
                  u.nombre AS uni,
                  ca.nombre AS car,
                  c.periodo,
                  c.nombre AS curso,
                  COALESCE(c.codigo,'') AS cod,
                  c.carrera_id,
                  ca.universidad_id
                FROM cursos c
                JOIN carreras ca ON ca.carrera_id=c.carrera_id
                JOIN universidades u ON u.universidad_id=ca.universidad_id
                WHERE
                  u.nombre LIKE ? COLLATE NOCASE OR
                  ca.nombre LIKE ? COLLATE NOCASE OR
                  c.periodo LIKE ? COLLATE NOCASE OR
                  c.nombre LIKE ? COLLATE NOCASE OR
                  COALESCE(c.codigo,'') LIKE ? COLLATE NOCASE
                ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
                """,
                (q, q, q, q, q),
                fetch=True,
            ) or []
        else:                                                       # Si no hay filtro
            rows = self._q(
                """
                SELECT
                  c.curso_id,
                  u.nombre AS uni,
                  ca.nombre AS car,
                  c.periodo,
                  c.nombre AS curso,
                  COALESCE(c.codigo,'') AS cod,
                  c.carrera_id,
                  ca.universidad_id
                FROM cursos c
                JOIN carreras ca ON ca.carrera_id=c.carrera_id
                JOIN universidades u ON u.universidad_id=ca.universidad_id
                ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
                """,
                fetch=True,
            ) or []

        for r in rows:                                              # Inserta filas
            self.tree.insert(
                "",
                "end",
                values=(int(r["curso_id"]), r["uni"], r["car"], r["periodo"], r["curso"], r["cod"]),
            )

        self._status(f"Vista: Cursos | Registros: {len(rows)}")     # Status

    def on_buscar(self):                                            # Acción buscar
        self._cargar_listado(self.var_buscar.get().strip())         # Recarga listado

    def on_ver_todo(self):                                          # Ver todo
        self.var_buscar.set("")                                     # Limpia filtro
        self._cargar_listado("")                                    # Recarga completo

    # =========================================================
    # Selección -> formulario
    # =========================================================
    def on_select(self):                                            # Al seleccionar curso
        sel = self.tree.selection()                                 # Selección
        if not sel:                                                 # Si no hay
            return                                                  # Sale

        vals = self.tree.item(sel[0], "values")                     # Valores
        if not vals:                                                # Si vacío
            return                                                  # Sale

        self.curso_sel_id = int(vals[0])                            # Guarda curso_id

        # Carga info completa del curso desde BD (para mapear IDs)
        row = self._q(
            """
            SELECT
              c.curso_id, c.periodo, c.nombre, COALESCE(c.codigo,'') AS codigo,
              ca.carrera_id, ca.nombre AS car_nombre,
              u.universidad_id, u.nombre AS uni_nombre
            FROM cursos c
            JOIN carreras ca ON ca.carrera_id=c.carrera_id
            JOIN universidades u ON u.universidad_id=ca.universidad_id
            WHERE c.curso_id=?
            """,
            (self.curso_sel_id,),
            fetch=True,
        )
        if not row:                                                 # Si no encuentra
            return                                                  # Sale

        r = row[0]                                                  # Primera fila

        # Ajusta universidad
        self._cargar_universidades()                                # Refresca combo universidades
        if r["uni_nombre"] in self._uni_display:                    # Si existe
            self.var_uni.set(r["uni_nombre"])                       # Selecciona en combo
            self.on_uni_change()                                    # Carga carreras de esa uni

        # Ajusta carrera
        if r["car_nombre"] in self._car_display:                    # Si existe
            self.var_car.set(r["car_nombre"])                       # Selecciona carrera

        # Carga campos simples
        self.var_periodo.set(r["periodo"])                          # Periodo
        self.var_nombre.set(r["nombre"])                            # Nombre curso
        self.var_codigo.set(r["codigo"])                            # Código

    # =========================================================
    # CRUD
    # =========================================================
    def on_nuevo(self):                                             # Nuevo registro
        self.curso_sel_id = None                                    # Sin selección
        self.on_limpiar()                                           # Limpia campos

    def on_limpiar(self):                                           # Limpia formulario
        self.var_periodo.set("")                                    # Limpia periodo
        self.var_nombre.set("")                                     # Limpia nombre
        self.var_codigo.set("")                                     # Limpia código

    def _validar_periodo(self, periodo: str) -> bool:               # Valida formato periodo YYYY-1 o YYYY-2
        p = (periodo or "").strip()                                 # Limpia
        if len(p) != 6:                                             # Debe ser 6 chars: 2026-1
            return False                                            # No válido
        if p[4] != "-":                                             # Debe tener guion
            return False                                            # No válido
        anio = p[:4]                                                # Año
        sem = p[5]                                                  # Semestre
        return anio.isdigit() and sem in ("1", "2")                 # Validación final

    def on_guardar(self):                                           # Guardar (INSERT/UPDATE)
        uni = self._uni_info_sel()                                  # Universidad seleccionada
        car = self._car_info_sel()                                  # Carrera seleccionada

        if not uni:                                                 # Validación
            messagebox.showwarning("Atención", "Seleccione una universidad.")  # Aviso
            return                                                  # Sale
        if not car:                                                 # Validación
            messagebox.showwarning("Atención", "Seleccione una carrera.")     # Aviso
            return                                                  # Sale

        periodo = self.var_periodo.get().strip()                    # Periodo
        nombre = self.var_nombre.get().strip()                      # Nombre
        codigo = self.var_codigo.get().strip()                      # Código

        if not self._validar_periodo(periodo):                      # Valida periodo
            messagebox.showwarning("Atención", "Periodo inválido. Use formato YYYY-1 o YYYY-2 (ej: 2026-1).")  # Aviso
            return                                                  # Sale
        if not nombre:                                              # Valida nombre
            messagebox.showwarning("Atención", "El nombre del curso es obligatorio.")  # Aviso
            return                                                  # Sale

        try:
            if self.curso_sel_id is None:                           # INSERT
                self._q(
                    "INSERT INTO cursos(carrera_id, periodo, nombre, codigo) VALUES(?,?,?,?)",
                    (car["id"], periodo, nombre, codigo),
                )
                messagebox.showinfo("OK", "Curso creado.")          # OK
            else:                                                   # UPDATE
                self._q(
                    "UPDATE cursos SET carrera_id=?, periodo=?, nombre=?, codigo=? WHERE curso_id=?",
                    (car["id"], periodo, nombre, codigo, self.curso_sel_id),
                )
                messagebox.showinfo("OK", "Curso actualizado.")     # OK

            self._cargar_listado(self.var_buscar.get().strip())     # Refresca listado
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    def on_eliminar(self):                                          # Eliminar curso
        if self.curso_sel_id is None:                               # Debe seleccionar
            messagebox.showwarning("Atención", "Seleccione un curso para eliminar.")  # Aviso
            return                                                  # Sale
        if not messagebox.askyesno("Confirmar", "¿Eliminar curso? (se eliminarán evaluaciones, inscripciones y notas asociadas)"):  # Confirm
            return                                                  # Cancel

        try:
            self._q("DELETE FROM cursos WHERE curso_id=?", (self.curso_sel_id,))  # Delete
            self.curso_sel_id = None                                # Limpia selección
            self.on_limpiar()                                       # Limpia formulario
            self._cargar_listado(self.var_buscar.get().strip())     # Refresca listado
            messagebox.showinfo("OK", "Curso eliminado.")           # OK
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    # =========================================================
    # Crear universidad/carrera rápido
    # =========================================================
    def on_nueva_universidad(self):                                 # Crea universidad desde UI
        nombre = simpledialog.askstring("Nueva universidad", "Nombre de la universidad:")  # Pide nombre
        if not nombre:                                              # Si cancel
            return                                                  # Sale
        nombre = nombre.strip()                                     # Limpia
        if not nombre:                                              # Si vacío
            return                                                  # Sale
        try:
            self._q("INSERT INTO universidades(nombre) VALUES(?)", (nombre,))  # Inserta
            self._cargar_universidades()                            # Refresca combo
            self.var_uni.set(nombre)                                # Selecciona nueva
            self.on_uni_change()                                    # Carga carreras
            messagebox.showinfo("OK", "Universidad creada.")        # OK
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    def on_nueva_carrera(self):                                     # Crea carrera desde UI
        uni = self._uni_info_sel()                                  # Universidad seleccionada
        if not uni:                                                 # Validación
            messagebox.showwarning("Atención", "Seleccione una universidad antes de crear una carrera.")  # Aviso
            return                                                  # Sale
        nombre = simpledialog.askstring("Nueva carrera", "Nombre de la carrera:")  # Pide nombre
        if not nombre:                                              # Cancel
            return                                                  # Sale
        nombre = nombre.strip()                                     # Limpia
        if not nombre:                                              # Vacío
            return                                                  # Sale
        try:
            self._q("INSERT INTO carreras(universidad_id, nombre) VALUES(?,?)", (uni["id"], nombre))  # Inserta
            self._cargar_carreras(uni["id"])                        # Refresca carreras
            self.var_car.set(nombre)                                # Selecciona nueva
            messagebox.showinfo("OK", "Carrera creada.")            # OK
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    # =========================================================
    # Exportaciones
    # =========================================================
    def _datos_export(self):                                        # Obtiene filas según filtro actual
        t = self.var_buscar.get().strip()                           # Texto filtro
        if t:                                                       # Con filtro
            q = f"%{t}%"                                           # Like
            rows = self._q(
                """
                SELECT u.nombre AS universidad, ca.nombre AS carrera, c.periodo, c.nombre AS curso, COALESCE(c.codigo,'') AS codigo
                FROM cursos c
                JOIN carreras ca ON ca.carrera_id=c.carrera_id
                JOIN universidades u ON u.universidad_id=ca.universidad_id
                WHERE u.nombre LIKE ? COLLATE NOCASE OR ca.nombre LIKE ? COLLATE NOCASE OR c.periodo LIKE ? COLLATE NOCASE OR c.nombre LIKE ? COLLATE NOCASE OR COALESCE(c.codigo,'') LIKE ? COLLATE NOCASE
                ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
                """,
                (q, q, q, q, q),
                fetch=True,
            ) or []
        else:                                                       # Sin filtro
            rows = self._q(
                """
                SELECT u.nombre AS universidad, ca.nombre AS carrera, c.periodo, c.nombre AS curso, COALESCE(c.codigo,'') AS codigo
                FROM cursos c
                JOIN carreras ca ON ca.carrera_id=c.carrera_id
                JOIN universidades u ON u.universidad_id=ca.universidad_id
                ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
                """,
                fetch=True,
            ) or []
        return rows                                                 # Retorna filas

    def on_exportar_excel(self):                                    # Exporta listado a Excel
        try:
            rows = self._datos_export()                             # Filas
            ruta = filedialog.asksaveasfilename(                    # Ruta
                title="Guardar Excel (Cursos)",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                            # Cancel
                return                                              # Sale

            wb = Workbook()                                         # Libro
            ws = wb.active                                          # Hoja
            ws.title = "Cursos"                                     # Nombre hoja

            ws.append(["Universidad", "Carrera", "Periodo", "Curso", "Codigo"])  # Cabecera
            for r in rows:                                          # Recorre filas
                ws.append([r["universidad"], r["carrera"], r["periodo"], r["curso"], r["codigo"]])  # Fila

            wb.save(ruta)                                           # Guarda
            messagebox.showinfo("OK", "Excel exportado correctamente.")  # OK
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    def on_exportar_pdf(self):                                      # Exporta listado a PDF
        try:
            rows = self._datos_export()                             # Filas
            ruta = filedialog.asksaveasfilename(                    # Ruta
                title="Guardar PDF (Cursos)",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                            # Cancel
                return                                              # Sale

            c = canvas.Canvas(ruta, pagesize=A4)                    # Canvas
            w, h = A4                                               # Dimensiones
            y = h - 60                                              # Cursor Y

            c.setFont("Helvetica-Bold", 13)                         # Fuente título
            c.drawString(40, y, "Listado de Cursos")                # Título
            y -= 20                                                 # Baja

            c.setFont("Helvetica-Bold", 9)                          # Cabecera
            c.drawString(40, y, "Universidad")                      # Col
            c.drawString(170, y, "Carrera")                         # Col
            c.drawString(330, y, "Periodo")                         # Col
            c.drawString(390, y, "Curso")                           # Col
            y -= 12                                                 # Baja

            c.setFont("Helvetica", 9)                               # Fuente normal
            for r in rows:                                          # Filas
                if y < 60:                                          # Salto página
                    c.showPage()                                    # Nueva página
                    y = h - 60                                      # Reinicia Y
                    c.setFont("Helvetica", 9)                       # Fuente
                c.drawString(40, y, str(r["universidad"])[:18])     # Uni recortada
                c.drawString(170, y, str(r["carrera"])[:22])        # Carrera recortada
                c.drawString(330, y, str(r["periodo"])[:10])        # Periodo
                c.drawString(390, y, str(r["curso"])[:30])          # Curso recortado
                y -= 12                                             # Baja

            c.save()                                                # Guarda PDF
            messagebox.showinfo("OK", "PDF exportado correctamente.")  # OK
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error

    # =========================================================
    # Importar Excel
    # - Crea universidad/carrera si no existen
    # - UPSERT curso por UNIQUE(carrera_id, periodo, nombre)
    # =========================================================
    def on_importar_excel(self):                                    # Importa cursos desde Excel
        ruta = filedialog.askopenfilename(                          # Selecciona archivo
            title="Importar Excel (Cursos)",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=".",
        )
        if not ruta:                                                # Cancel
            return                                                  # Sale

        try:
            wb = load_workbook(ruta)                                # Abre libro
            ws = wb.active                                          # Hoja activa

            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Cabecera
            headers = [str(h or "").strip().casefold() for h in header]           # Normaliza

            def hidx(*names: str) -> int:                           # Busca índice de columna
                for n in names:                                     # Recorre alias
                    k = n.casefold()                                # Normaliza
                    if k in headers:                                # Si existe
                        return headers.index(k)                     # Índice
                return -1                                           # No encontrado

            i_uni = hidx("universidad")                             # Col universidad
            i_car = hidx("carrera")                                 # Col carrera
            i_per = hidx("periodo", "período")                      # Col periodo
            i_cur = hidx("curso", "nombrecurso", "nombre curso")    # Col curso
            i_cod = hidx("codigo", "código")                        # Col código (opcional)

            if i_uni == -1 or i_car == -1 or i_per == -1 or i_cur == -1:  # Validación
                raise ValueError("Excel debe incluir: Universidad, Carrera, Periodo, Curso (y opcional Codigo).")  # Error

            conn = obtener_conexion()                               # Conexión única para performance
            try:
                cur = conn.cursor()                                 # Cursor

                total = 0                                           # Total procesadas
                ok = 0                                              # OK
                errores = []                                        # Errores

                for fila_n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Datos
                    total += 1                                      # Cuenta
                    try:
                        uni = str(row[i_uni] or "").strip()         # Universidad
                        car = str(row[i_car] or "").strip()         # Carrera
                        per = str(row[i_per] or "").strip()         # Periodo
                        cur_nom = str(row[i_cur] or "").strip()     # Curso
                        cod = ""                                    # Código por defecto
                        if i_cod != -1:                             # Si existe columna código
                            cod = str(row[i_cod] or "").strip()     # Lee código

                        if not uni or not car or not per or not cur_nom:  # Validación
                            raise ValueError("Faltan datos obligatorios (Universidad/Carrera/Periodo/Curso).")  # Error

                        if not self._validar_periodo(per):          # Valida periodo
                            raise ValueError("Periodo inválido (use YYYY-1 o YYYY-2).")  # Error

                        # 1) Universidad: buscar/crear
                        cur.execute("SELECT universidad_id FROM universidades WHERE nombre=?", (uni,))  # Busca
                        r = cur.fetchone()                           # Fila
                        if r:                                        # Si existe
                            uni_id = int(r[0])                       # ID
                        else:                                        # Si no existe
                            cur.execute("INSERT INTO universidades(nombre) VALUES(?)", (uni,))  # Crea
                            uni_id = int(cur.lastrowid)              # Nuevo ID

                        # 2) Carrera: buscar/crear (UNIQUE(universidad_id, nombre))
                        cur.execute("SELECT carrera_id FROM carreras WHERE universidad_id=? AND nombre=?", (uni_id, car))  # Busca
                        r = cur.fetchone()                           # Fila
                        if r:                                        # Existe
                            car_id = int(r[0])                       # ID
                        else:                                        # No existe
                            cur.execute("INSERT INTO carreras(universidad_id, nombre) VALUES(?,?)", (uni_id, car))  # Crea
                            car_id = int(cur.lastrowid)              # ID

                        # 3) Curso: UPSERT por UNIQUE(carrera_id, periodo, nombre)
                        cur.execute(
                            """
                            INSERT INTO cursos(carrera_id, periodo, nombre, codigo)
                            VALUES(?,?,?,?)
                            ON CONFLICT(carrera_id, periodo, nombre)
                            DO UPDATE SET codigo=excluded.codigo
                            """,
                            (car_id, per, cur_nom, cod),
                        )

                        ok += 1                                     # OK
                    except Exception as e:
                        errores.append(f"Fila {fila_n}: {e}")        # Guarda error

                conn.commit()                                       # Confirma todo
            finally:
                conn.close()                                        # Cierra conexión

            self._cargar_universidades()                             # Refresca combos
            self._cargar_listado(self.var_buscar.get().strip())      # Refresca listado

            msg = f"Importación finalizada.\nProcesadas: {total}\nOK: {ok}\nErrores: {len(errores)}"  # Resumen
            if errores:                                              # Si hubo errores
                msg += "\n\nPrimeros errores:\n" + "\n".join(errores[:10])  # Muestra 10
            messagebox.showinfo("Resultado", msg)                    # Muestra resumen

        except Exception as e:
            messagebox.showerror("Error", str(e))                    # Error general
