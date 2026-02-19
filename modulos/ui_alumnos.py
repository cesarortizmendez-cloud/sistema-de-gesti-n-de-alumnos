# ============================================  # Separador visual
# modulos/ui_alumnos.py                          # Página "Alumnos" (UI única, responsiva)
# - Izquierda: listado (Treeview)                # Visualiza alumnos
# - Derecha: formulario CRUD                     # Crear/editar/eliminar alumno
# - Barra superior compacta: buscar + export/import # Acciones rápidas
# - Incluye migración suave de columnas faltantes # Evita error rut_normalizado
# - Exportar Excel/PDF e Importar Excel           # Requerimientos actividad
# ============================================  # Separador visual

import tkinter as tk                                               # Tkinter base                                     # noqa: E501
from tkinter import ttk, messagebox, filedialog                    # Widgets + diálogos                               # noqa: E501
from datetime import datetime                                      # Para periodo por defecto                          # noqa: E501

from openpyxl import Workbook                                      # Exportar Excel                                   # noqa: E501
from openpyxl import load_workbook                                 # Importar Excel                                   # noqa: E501

from reportlab.pdfgen import canvas                                # Crear PDF                                        # noqa: E501
from reportlab.lib.pagesizes import A4                             # Tamaño A4                                        # noqa: E501

from .bd_sqlite import obtener_conexion                            # Conexión SQLite                                  # noqa: E501
from .config import carpeta_exports                                # Carpeta exports                                  # noqa: E501


class PaginaAlumnos(ttk.Frame):                                    # Página Alumnos (Frame embebido)                  # noqa: E501
    def __init__(self, master):                                    # Constructor                                      # noqa: E501
        super().__init__(master)                                   # Inicializa Frame                                 # noqa: E501

        self.alumno_sel_id = None                                  # ID alumno seleccionado                           # noqa: E501

        self._universidades = []                                   # Lista de universidades                           # noqa: E501
        self._carreras = []                                        # Lista de carreras de la uni seleccionada         # noqa: E501
        self._uni_display = []                                     # Strings para combobox universidad                # noqa: E501
        self._car_display = []                                     # Strings para combobox carrera                    # noqa: E501

        self._asegurar_schema_alumnos()                            # ✅ Migración suave si BD vieja                    # noqa: E501

        self._crear_ui()                                           # Construye UI                                     # noqa: E501
        self._cargar_universidades()                               # Carga universidades                              # noqa: E501
        self._cargar_listado("")                                   # Carga listado completo                            # noqa: E501

    def on_show(self):                                             # Hook al mostrar página                            # noqa: E501
        self._asegurar_schema_alumnos()                            # Re-chequea schema (por seguridad)                 # noqa: E501
        self._cargar_universidades()                               # Refresca combos                                   # noqa: E501
        self._cargar_listado(self.var_buscar.get().strip())        # Refresca listado                                  # noqa: E501

    # =========================================================
    # Statusbar (si ventana principal tiene set_status)
    # =========================================================
    def _status(self, texto: str):                                 # Envía texto a barra de estado                     # noqa: E501
        top = self.winfo_toplevel()                                # Ventana principal                                 # noqa: E501
        if hasattr(top, "set_status") and callable(getattr(top, "set_status")):  # Si existe set_status          # noqa: E501
            top.set_status(texto)                                  # Actualiza estado                                  # noqa: E501

    # =========================================================
    # Helper SQL simple
    # =========================================================
    def _q(self, sql: str, params: tuple = (), fetch: bool = False):  # Ejecuta SQL con parámetros                       # noqa: E501
        conn = obtener_conexion()                                  # Abre conexión                                     # noqa: E501
        try:                                                       # Bloque seguro                                     # noqa: E501
            cur = conn.cursor()                                    # Cursor                                            # noqa: E501
            cur.execute(sql, params)                               # Ejecuta                                            # noqa: E501
            if fetch:                                              # Si requiere resultados                             # noqa: E501
                return cur.fetchall()                              # Devuelve filas                                     # noqa: E501
            conn.commit()                                          # Confirma cambios                                   # noqa: E501
            return None                                            # None                                               # noqa: E501
        finally:                                                   # Siempre                                             # noqa: E501
            conn.close()                                           # Cierra conexión                                    # noqa: E501

    # =========================================================
    # ✅ Migración suave: arregla BD existente si faltan columnas
    # =========================================================
    def _asegurar_schema_alumnos(self):                            # Asegura columnas requeridas                         # noqa: E501
        conn = obtener_conexion()                                  # Abre conexión                                       # noqa: E501
        try:                                                       # Bloque seguro                                       # noqa: E501
            cur = conn.cursor()                                    # Cursor                                              # noqa: E501

            # Verifica que exista la tabla alumnos
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='alumnos'")  # Busca tabla      # noqa: E501
            if cur.fetchone() is None:                             # Si no existe                                        # noqa: E501
                return                                             # Sale (bd_sqlite debe crearla)                       # noqa: E501

            # Lista columnas actuales de alumnos
            cur.execute("PRAGMA table_info(alumnos)")              # Información de columnas                             # noqa: E501
            cols = [r["name"] for r in cur.fetchall()]             # Extrae nombres                                      # noqa: E501

            # Periodo por defecto (ej: 2026-1)
            periodo_default = f"{datetime.now().year}-1"           # Periodo default                                     # noqa: E501

            # Si falta rut_normalizado => la agrega
            if "rut_normalizado" not in cols:                      # Si no existe                                        # noqa: E501
                cur.execute("ALTER TABLE alumnos ADD COLUMN rut_normalizado TEXT")  # Agrega columna            # noqa: E501
                # Rellena rut_normalizado con rut limpiado (básico)
                cur.execute(
                    """
                    UPDATE alumnos
                    SET rut_normalizado = UPPER(REPLACE(REPLACE(REPLACE(COALESCE(rut,''),'.',''),'-',''),' ',''))
                    WHERE rut_normalizado IS NULL OR TRIM(rut_normalizado)=''
                    """
                )                                                  # Normaliza rut existente                              # noqa: E501

            # Si falta nombre_busqueda => la agrega (sirve para buscar rápido)
            if "nombre_busqueda" not in cols:                      # Si no existe                                        # noqa: E501
                cur.execute("ALTER TABLE alumnos ADD COLUMN nombre_busqueda TEXT")  # Agrega columna            # noqa: E501
                cur.execute(
                    """
                    UPDATE alumnos
                    SET nombre_busqueda = LOWER(TRIM(COALESCE(apellidos,'') || ' ' || COALESCE(nombres,'') || ' ' || COALESCE(rut,'')))
                    WHERE nombre_busqueda IS NULL OR TRIM(nombre_busqueda)=''
                    """
                )                                                  # Rellena nombre_busqueda                              # noqa: E501

            # Si falta periodo => la agrega
            if "periodo" not in cols:                              # Si no existe                                        # noqa: E501
                cur.execute("ALTER TABLE alumnos ADD COLUMN periodo TEXT")  # Agrega columna                              # noqa: E501
                cur.execute(
                    "UPDATE alumnos SET periodo=? WHERE periodo IS NULL OR TRIM(periodo)=''",
                    (periodo_default,),
                )                                                  # Rellena periodo                                     # noqa: E501

            # Si falta tipo_alumno => la agrega
            if "tipo_alumno" not in cols:                          # Si no existe                                        # noqa: E501
                cur.execute("ALTER TABLE alumnos ADD COLUMN tipo_alumno TEXT")  # Agrega columna                            # noqa: E501
                cur.execute(
                    "UPDATE alumnos SET tipo_alumno='Pregrado' WHERE tipo_alumno IS NULL OR TRIM(tipo_alumno)=''"
                )                                                  # Rellena tipo                                        # noqa: E501

            # Si falta estado => la agrega
            if "estado" not in cols:                               # Si no existe                                        # noqa: E501
                cur.execute("ALTER TABLE alumnos ADD COLUMN estado INTEGER")  # Agrega columna                              # noqa: E501
                cur.execute(
                    "UPDATE alumnos SET estado=1 WHERE estado IS NULL"
                )                                                  # Activa por defecto                                  # noqa: E501

            conn.commit()                                          # Confirma migraciones                                 # noqa: E501
        finally:                                                   # Siempre                                              # noqa: E501
            conn.close()                                           # Cierra conexión                                      # noqa: E501

    # =========================================================
    # Normalización de RUT (básica) para evitar duplicados
    # =========================================================
    def _rut_norm(self, rut: str) -> str:                           # Normaliza rut (numero+dv sin símbolos)             # noqa: E501
        t = (rut or "").strip().upper()                             # Limpia y mayúsculas                                 # noqa: E501
        t = t.replace(".", "").replace(" ", "")                     # Quita puntos y espacios                              # noqa: E501
        if "-" not in t and len(t) >= 2:                            # Si no trae guion                                     # noqa: E501
            t = t[:-1] + "-" + t[-1]                                # Inserta guion                                        # noqa: E501
        parts = t.split("-")                                        # Separa                                                # noqa: E501
        if len(parts) != 2:                                         # Debe tener 2 partes                                  # noqa: E501
            return ""                                               # Invalida                                             # noqa: E501
        num = parts[0]                                              # Parte numérica                                       # noqa: E501
        dv = parts[1]                                               # Dígito verificador                                   # noqa: E501
        if not num.isdigit():                                       # Valida numérico                                      # noqa: E501
            return ""                                               # Invalida                                             # noqa: E501
        return f"{int(num)}{dv}"                                    # Devuelve normalizado                                 # noqa: E501

    def _periodo_valido(self, p: str) -> bool:                       # Valida YYYY-1 / YYYY-2                               # noqa: E501
        p = (p or "").strip()                                       # Limpia                                                # noqa: E501
        if len(p) != 6:                                             # Debe ser 6                                            # noqa: E501
            return False                                            # No                                                    # noqa: E501
        if p[4] != "-":                                             # Debe tener guion                                      # noqa: E501
            return False                                            # No                                                    # noqa: E501
        return p[:4].isdigit() and p[5] in ("1", "2")               # Año y semestre validos                                # noqa: E501

    # =========================================================
    # UI
    # =========================================================
    def _crear_ui(self):                                            # Construye interfaz                                   # noqa: E501
        self.columnconfigure(0, weight=1)                           # Crece ancho                                            # noqa: E501
        self.rowconfigure(2, weight=1)                              # ✅ Solo contenido crece alto                           # noqa: E501

        ttk.Label(self, text="Alumnos", font=("Segoe UI", 16, "bold")).grid(  # Título                                         # noqa: E501
            row=0, column=0, sticky="w", padx=12, pady=(12, 6)
        )

        # Barra superior compacta
        barra = ttk.Frame(self)                                     # Frame barra                                            # noqa: E501
        barra.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        barra.columnconfigure(1, weight=1)                          # Buscar se expande                                      # noqa: E501

        ttk.Label(barra, text="Buscar:").grid(row=0, column=0, sticky="w")  # Label buscar                                     # noqa: E501
        self.var_buscar = tk.StringVar()                            # Variable buscar                                        # noqa: E501
        ent = ttk.Entry(barra, textvariable=self.var_buscar)        # Entry buscar                                           # noqa: E501
        ent.grid(row=0, column=1, sticky="ew", padx=(8, 10))         # Ubicación                                              # noqa: E501
        ent.bind("<Return>", lambda e: self.on_buscar())            # Enter -> buscar                                        # noqa: E501

        ttk.Button(barra, text="Aplicar", command=self.on_buscar).grid(row=0, column=2, padx=4)      # Botón aplicar      # noqa: E501
        ttk.Button(barra, text="Ver todo", command=self.on_ver_todo).grid(row=0, column=3, padx=4)   # Botón ver todo     # noqa: E501

        ttk.Separator(barra, orient="vertical").grid(row=0, column=4, sticky="ns", padx=10)          # Separador          # noqa: E501
        ttk.Button(barra, text="Exportar Excel", command=self.on_exportar_excel).grid(row=0, column=5, padx=4)  # Excel    # noqa: E501
        ttk.Button(barra, text="Exportar PDF", command=self.on_exportar_pdf).grid(row=0, column=6, padx=4)      # PDF      # noqa: E501
        ttk.Button(barra, text="Importar Excel", command=self.on_importar_excel).grid(row=0, column=7, padx=4)  # Import   # noqa: E501

        # Contenido principal (izq listado / der formulario)
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)     # PanedWindow                                            # noqa: E501
        self.paned.grid(row=2, column=0, sticky="nsew", padx=12, pady=12)

        self.panel_izq = ttk.Frame(self.paned)                      # Panel izquierdo                                        # noqa: E501
        self.panel_der = ttk.Frame(self.paned)                      # Panel derecho                                          # noqa: E501

        self.panel_izq.columnconfigure(0, weight=1)                 # Tabla crece                                            # noqa: E501
        self.panel_izq.rowconfigure(0, weight=1)                    # Tabla crece                                            # noqa: E501

        self.panel_der.columnconfigure(0, weight=1)                 # Form crece                                             # noqa: E501
        self.panel_der.rowconfigure(0, weight=1)                    # Form crece                                             # noqa: E501

        self.paned.add(self.panel_izq, weight=3)                    # Izq más ancho                                          # noqa: E501
        self.paned.add(self.panel_der, weight=2)                    # Der                                                    # noqa: E501

        try:                                                        # Config mínimos panes                                   # noqa: E501
            self.paned.paneconfigure(self.panel_izq, minsize=650)    # Min izq                                                # noqa: E501
            self.paned.paneconfigure(self.panel_der, minsize=450)    # Min der                                                # noqa: E501
        except Exception:                                           # Si no soporta                                          # noqa: E501
            pass                                                     # No rompe                                               # noqa: E501

        # -------- Izquierda: Treeview alumnos --------
        self.tree = ttk.Treeview(                                   # Tabla alumnos                                          # noqa: E501
            self.panel_izq,
            columns=("id", "rut", "alumno", "uni", "car", "per", "tipo", "est"),
            show="headings",
            selectmode="browse",
        )

        for col, text in [                                          # Define headers                                         # noqa: E501
            ("id", "ID"),
            ("rut", "RUT"),
            ("alumno", "Alumno"),
            ("uni", "Universidad"),
            ("car", "Carrera"),
            ("per", "Periodo"),
            ("tipo", "Tipo"),
            ("est", "Estado"),
        ]:
            self.tree.heading(col, text=text)                       # Asigna encabezado                                      # noqa: E501

        self.tree.column("id", width=60, anchor="center")           # Col ID                                                 # noqa: E501
        self.tree.column("rut", width=120, anchor="w")              # Col RUT                                                # noqa: E501
        self.tree.column("alumno", width=220, anchor="w")           # Col Alumno                                             # noqa: E501
        self.tree.column("uni", width=180, anchor="w")              # Col Uni                                                # noqa: E501
        self.tree.column("car", width=180, anchor="w")              # Col Carrera                                            # noqa: E501
        self.tree.column("per", width=90, anchor="center")          # Col Periodo                                            # noqa: E501
        self.tree.column("tipo", width=110, anchor="center")        # Col Tipo                                               # noqa: E501
        self.tree.column("est", width=80, anchor="center")          # Col Estado                                             # noqa: E501

        sb = ttk.Scrollbar(self.panel_izq, orient="vertical", command=self.tree.yview)  # Scroll                                      # noqa: E501
        self.tree.configure(yscrollcommand=sb.set)                  # Enlaza scroll                                          # noqa: E501

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 6))  # Ubicación tree                                         # noqa: E501
        sb.grid(row=0, column=1, sticky="ns")                       # Ubicación scroll                                       # noqa: E501

        self.tree.bind("<<TreeviewSelect>>", lambda e: self.on_select())  # Selección -> form                                  # noqa: E501

        # -------- Derecha: Formulario --------
        box = ttk.LabelFrame(self.panel_der, text="Formulario")     # Marco formulario                                      # noqa: E501
        box.grid(row=0, column=0, sticky="nsew")                    # Ubicación                                              # noqa: E501
        box.columnconfigure(1, weight=1)                            # Columna entrada crece                                  # noqa: E501

        # Variables
        self.var_tipo = tk.StringVar(value="Pregrado")              # Tipo alumno                                            # noqa: E501
        self.var_rut = tk.StringVar()                               # RUT                                                    # noqa: E501
        self.var_nombres = tk.StringVar()                           # Nombres                                                # noqa: E501
        self.var_apellidos = tk.StringVar()                         # Apellidos                                              # noqa: E501
        self.var_email = tk.StringVar()                             # Email                                                  # noqa: E501
        self.var_telefono = tk.StringVar()                          # Teléfono                                               # noqa: E501
        self.var_uni = tk.StringVar()                               # Universidad (texto)                                    # noqa: E501
        self.var_car = tk.StringVar()                               # Carrera (texto)                                        # noqa: E501
        self.var_periodo = tk.StringVar(value=f"{datetime.now().year}-1")  # Periodo default                                   # noqa: E501
        self.var_activo = tk.IntVar(value=1)                        # Estado 1=activo                                        # noqa: E501

        # Tipo alumno
        ttk.Label(box, text="Tipo alumno:").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 6))  # Label          # noqa: E501
        cmb_tipo = ttk.Combobox(box, textvariable=self.var_tipo, state="readonly",
                                values=["Pregrado", "Postgrado", "Intercambio"])  # Combo tipo                                      # noqa: E501
        cmb_tipo.grid(row=0, column=1, sticky="ew", padx=10, pady=(10, 6))      # Ubicación combo                                      # noqa: E501

        # RUT
        ttk.Label(box, text="RUT:").grid(row=1, column=0, sticky="w", padx=10, pady=6)                # Label          # noqa: E501
        ttk.Entry(box, textvariable=self.var_rut).grid(row=1, column=1, sticky="ew", padx=10, pady=6) # Entry          # noqa: E501

        # Nombres
        ttk.Label(box, text="Nombres:").grid(row=2, column=0, sticky="w", padx=10, pady=6)            # Label          # noqa: E501
        ttk.Entry(box, textvariable=self.var_nombres).grid(row=2, column=1, sticky="ew", padx=10, pady=6)  # Entry      # noqa: E501

        # Apellidos
        ttk.Label(box, text="Apellidos:").grid(row=3, column=0, sticky="w", padx=10, pady=6)          # Label          # noqa: E501
        ttk.Entry(box, textvariable=self.var_apellidos).grid(row=3, column=1, sticky="ew", padx=10, pady=6) # Entry       # noqa: E501

        # Email
        ttk.Label(box, text="Email:").grid(row=4, column=0, sticky="w", padx=10, pady=6)              # Label          # noqa: E501
        ttk.Entry(box, textvariable=self.var_email).grid(row=4, column=1, sticky="ew", padx=10, pady=6)     # Entry       # noqa: E501

        # Teléfono
        ttk.Label(box, text="Teléfono:").grid(row=5, column=0, sticky="w", padx=10, pady=6)           # Label          # noqa: E501
        ttk.Entry(box, textvariable=self.var_telefono).grid(row=5, column=1, sticky="ew", padx=10, pady=6)  # Entry       # noqa: E501

        # Universidad
        ttk.Label(box, text="Universidad:").grid(row=6, column=0, sticky="w", padx=10, pady=6)        # Label          # noqa: E501
        fila_uni = ttk.Frame(box)                                                                    # Frame combo+ +  # noqa: E501
        fila_uni.grid(row=6, column=1, sticky="ew", padx=10, pady=6)                                  # Ubicación       # noqa: E501
        fila_uni.columnconfigure(0, weight=1)                                                        # Combo crece      # noqa: E501
        self.cmb_uni = ttk.Combobox(fila_uni, textvariable=self.var_uni, state="readonly")           # Combo uni        # noqa: E501
        self.cmb_uni.grid(row=0, column=0, sticky="ew")                                              # Ubicación        # noqa: E501
        self.cmb_uni.bind("<<ComboboxSelected>>", lambda e: self.on_uni_change())                    # Evento           # noqa: E501
        ttk.Button(fila_uni, text="+", width=3, command=self.on_nueva_universidad).grid(row=0, column=1, padx=(6, 0))  # +  # noqa: E501

        # Carrera
        ttk.Label(box, text="Carrera:").grid(row=7, column=0, sticky="w", padx=10, pady=6)            # Label          # noqa: E501
        fila_car = ttk.Frame(box)                                                                    # Frame combo+ +  # noqa: E501
        fila_car.grid(row=7, column=1, sticky="ew", padx=10, pady=6)                                  # Ubicación       # noqa: E501
        fila_car.columnconfigure(0, weight=1)                                                        # Combo crece      # noqa: E501
        self.cmb_car = ttk.Combobox(fila_car, textvariable=self.var_car, state="readonly")           # Combo carrera    # noqa: E501
        self.cmb_car.grid(row=0, column=0, sticky="ew")                                              # Ubicación        # noqa: E501
        ttk.Button(fila_car, text="+", width=3, command=self.on_nueva_carrera).grid(row=0, column=1, padx=(6, 0))       # +  # noqa: E501

        # Periodo
        ttk.Label(box, text="Periodo (YYYY-1 / YYYY-2):").grid(row=8, column=0, sticky="w", padx=10, pady=6)           # Label # noqa: E501
        ttk.Entry(box, textvariable=self.var_periodo).grid(row=8, column=1, sticky="ew", padx=10, pady=6)              # Entry # noqa: E501

        # Activo
        ttk.Checkbutton(box, text="Activo", variable=self.var_activo).grid(row=9, column=1, sticky="w", padx=10, pady=(6, 6))  # Check  # noqa: E501

        # Botones CRUD
        btns = ttk.Frame(box)                                                                     # Frame botones      # noqa: E501
        btns.grid(row=10, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 10))            # Ubicación          # noqa: E501
        ttk.Button(btns, text="Nuevo", command=self.on_nuevo).pack(side="left", padx=6)           # Nuevo              # noqa: E501
        ttk.Button(btns, text="Guardar", command=self.on_guardar).pack(side="left", padx=6)      # Guardar            # noqa: E501
        ttk.Button(btns, text="Eliminar", command=self.on_eliminar).pack(side="left", padx=6)    # Eliminar           # noqa: E501
        ttk.Button(btns, text="Limpiar", command=self.on_limpiar).pack(side="left", padx=6)      # Limpiar            # noqa: E501

        ttk.Separator(box, orient="horizontal").grid(row=11, column=0, columnspan=2, sticky="ew", padx=10, pady=10)    # Sep # noqa: E501
        ttk.Label(
            box,
            text="Importar Excel: TipoAlumno | RUT | Nombres | Apellidos | Email | Telefono | Universidad | Carrera | Periodo | Estado(0/1)",
            foreground="gray",
            wraplength=380,
        ).grid(row=12, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 10))                                       # Tip # noqa: E501

    # =========================================================
    # Cargar combos (universidades/carreras)
    # =========================================================
    def _cargar_universidades(self):                                  # Carga universidades                               # noqa: E501
        rows = self._q("SELECT universidad_id, nombre FROM universidades ORDER BY nombre", fetch=True) or []  # Query # noqa: E501
        self._universidades = [{"id": int(r["universidad_id"]), "nombre": r["nombre"]} for r in rows]         # Lista # noqa: E501
        self._uni_display = [u["nombre"] for u in self._universidades]                                         # Display # noqa: E501
        self.cmb_uni["values"] = self._uni_display                                                   # Set combo # noqa: E501
        if self._uni_display:                                                                       # Si hay    # noqa: E501
            if self.var_uni.get().strip() not in self._uni_display:                                   # Si invál  # noqa: E501
                self.var_uni.set(self._uni_display[0])                                               # Elige 1°  # noqa: E501
            self.on_uni_change()                                                                     # Carga car # noqa: E501
        else:                                                                                        # Si no hay # noqa: E501
            self.var_uni.set("")                                                                     # Limpia    # noqa: E501
            self.var_car.set("")                                                                     # Limpia    # noqa: E501
            self.cmb_car["values"] = []                                                              # Limpia    # noqa: E501

    def _uni_sel(self):                                             # Retorna uni seleccionada (dict)                    # noqa: E501
        txt = self.var_uni.get().strip()                             # Texto                                              # noqa: E501
        if not txt or txt not in self._uni_display:                  # Validación                                         # noqa: E501
            return None                                              # None                                               # noqa: E501
        return self._universidades[self._uni_display.index(txt)]     # Dict uni                                           # noqa: E501

    def _cargar_carreras(self, uni_id: int):                         # Carga carreras por universidad                      # noqa: E501
        rows = self._q("SELECT carrera_id, nombre FROM carreras WHERE universidad_id=? ORDER BY nombre", (uni_id,), fetch=True) or []  # Query # noqa: E501
        self._carreras = [{"id": int(r["carrera_id"]), "nombre": r["nombre"]} for r in rows]          # Lista # noqa: E501
        self._car_display = [c["nombre"] for c in self._carreras]                                       # Display # noqa: E501
        self.cmb_car["values"] = self._car_display                                                     # Set combo # noqa: E501
        if self._car_display:                                                                          # Si hay    # noqa: E501
            if self.var_car.get().strip() not in self._car_display:                                    # Si invál  # noqa: E501
                self.var_car.set(self._car_display[0])                                                 # Elige 1°  # noqa: E501
        else:                                                                                           # Si no     # noqa: E501
            self.var_car.set("")                                                                       # Limpia    # noqa: E501

    def _car_sel(self):                                            # Retorna carrera seleccionada (dict)                # noqa: E501
        txt = self.var_car.get().strip()                            # Texto                                              # noqa: E501
        if not txt or txt not in self._car_display:                 # Validación                                         # noqa: E501
            return None                                             # None                                               # noqa: E501
        return self._carreras[self._car_display.index(txt)]         # Dict carrera                                       # noqa: E501

    def on_uni_change(self):                                       # Evento cambio universidad                           # noqa: E501
        uni = self._uni_sel()                                      # Dict uni                                           # noqa: E501
        if not uni:                                                # Si no hay                                          # noqa: E501
            self.cmb_car["values"] = []                            # Limpia carreras                                    # noqa: E501
            self.var_car.set("")                                   # Limpia carrera                                     # noqa: E501
            return                                                 # Sale                                               # noqa: E501
        self._cargar_carreras(uni["id"])                           # Carga carreras                                     # noqa: E501

    # =========================================================
    # Listado / búsqueda
    # =========================================================
    def _cargar_listado(self, texto: str):                          # Carga listado alumnos                               # noqa: E501
        for i in self.tree.get_children():                          # Limpia tree                                        # noqa: E501
            self.tree.delete(i)                                     # Borra                                              # noqa: E501

        t = (texto or "").strip()                                   # Texto filtro                                       # noqa: E501

        if t:                                                       # Con filtro                                         # noqa: E501
            q = f"%{t}%"                                           # Like                                               # noqa: E501
            rows = self._q(
                """
                SELECT
                  a.alumno_id,
                  a.rut,
                  a.nombres,
                  a.apellidos,
                  COALESCE(u.nombre,'') AS uni,
                  COALESCE(ca.nombre,'') AS car,
                  COALESCE(a.periodo,'') AS periodo,
                  COALESCE(a.tipo_alumno,'') AS tipo,
                  COALESCE(a.estado,1) AS estado
                FROM alumnos a
                LEFT JOIN universidades u ON u.universidad_id=a.universidad_id
                LEFT JOIN carreras ca ON ca.carrera_id=a.carrera_id
                WHERE
                  COALESCE(a.rut,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.nombres,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.apellidos,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.email,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.nombre_busqueda,'') LIKE ? COLLATE NOCASE
                ORDER BY a.apellidos, a.nombres
                """,
                (q, q, q, q, q),
                fetch=True,
            ) or []
        else:                                                       # Sin filtro                                         # noqa: E501
            rows = self._q(
                """
                SELECT
                  a.alumno_id,
                  a.rut,
                  a.nombres,
                  a.apellidos,
                  COALESCE(u.nombre,'') AS uni,
                  COALESCE(ca.nombre,'') AS car,
                  COALESCE(a.periodo,'') AS periodo,
                  COALESCE(a.tipo_alumno,'') AS tipo,
                  COALESCE(a.estado,1) AS estado
                FROM alumnos a
                LEFT JOIN universidades u ON u.universidad_id=a.universidad_id
                LEFT JOIN carreras ca ON ca.carrera_id=a.carrera_id
                ORDER BY a.apellidos, a.nombres
                """,
                fetch=True,
            ) or []

        for r in rows:                                              # Inserta filas                                      # noqa: E501
            alumno = f"{r['apellidos']} {r['nombres']}".strip()     # Nombre completo                                     # noqa: E501
            est = "Activo" if int(r["estado"] or 1) == 1 else "Inactivo"  # Estado texto                                  # noqa: E501
            self.tree.insert(
                "",
                "end",
                values=(int(r["alumno_id"]), r["rut"], alumno, r["uni"], r["car"], r["periodo"], r["tipo"], est),
            )

        self._status(f"Vista: Alumnos | Registros: {len(rows)}")     # Status                                              # noqa: E501

    def on_buscar(self):                                            # Acción buscar                                       # noqa: E501
        self._cargar_listado(self.var_buscar.get().strip())         # Recarga listado                                     # noqa: E501

    def on_ver_todo(self):                                          # Ver todo                                            # noqa: E501
        self.var_buscar.set("")                                     # Limpia filtro                                       # noqa: E501
        self._cargar_listado("")                                    # Recarga completo                                    # noqa: E501

    # =========================================================
    # Selección -> formulario
    # =========================================================
    def on_select(self):                                            # Selección en Treeview                               # noqa: E501
        sel = self.tree.selection()                                 # Selección                                            # noqa: E501
        if not sel:                                                 # Si no hay                                            # noqa: E501
            return                                                  # Sale                                                 # noqa: E501
        vals = self.tree.item(sel[0], "values")                     # Valores                                              # noqa: E501
        if not vals:                                                # Si vacío                                             # noqa: E501
            return                                                  # Sale                                                 # noqa: E501

        self.alumno_sel_id = int(vals[0])                           # Guarda alumno_id                                    # noqa: E501

        row = self._q(                                              # Carga alumno desde BD                               # noqa: E501
            """
            SELECT
              a.*,
              COALESCE(u.nombre,'') AS uni_nombre,
              COALESCE(ca.nombre,'') AS car_nombre
            FROM alumnos a
            LEFT JOIN universidades u ON u.universidad_id=a.universidad_id
            LEFT JOIN carreras ca ON ca.carrera_id=a.carrera_id
            WHERE a.alumno_id=?
            """,
            (self.alumno_sel_id,),
            fetch=True,
        )
        if not row:                                                 # Si no hay                                            # noqa: E501
            return                                                  # Sale                                                 # noqa: E501
        r = row[0]                                                  # Fila                                                  # noqa: E501

        self.var_tipo.set(r.get("tipo_alumno") or "Pregrado")       # Tipo                                                  # noqa: E501
        self.var_rut.set(r.get("rut") or "")                        # RUT                                                   # noqa: E501
        self.var_nombres.set(r.get("nombres") or "")                # Nombres                                               # noqa: E501
        self.var_apellidos.set(r.get("apellidos") or "")            # Apellidos                                             # noqa: E501
        self.var_email.set(r.get("email") or "")                    # Email                                                 # noqa: E501
        self.var_telefono.set(r.get("telefono") or "")              # Teléfono                                              # noqa: E501
        self.var_periodo.set(r.get("periodo") or f"{datetime.now().year}-1")  # Periodo                                        # noqa: E501
        self.var_activo.set(int(r.get("estado") or 1))              # Activo                                                # noqa: E501

        # Selecciona universidad/carrera en combos
        self._cargar_universidades()                                # Refresca combos                                      # noqa: E501
        uni_nombre = r.get("uni_nombre") or ""                      # Uni texto                                              # noqa: E501
        car_nombre = r.get("car_nombre") or ""                      # Carrera texto                                          # noqa: E501
        if uni_nombre in self._uni_display:                         # Si existe                                              # noqa: E501
            self.var_uni.set(uni_nombre)                            # Selecciona                                             # noqa: E501
            self.on_uni_change()                                    # Carga carreras                                         # noqa: E501
        if car_nombre in self._car_display:                         # Si existe                                              # noqa: E501
            self.var_car.set(car_nombre)                            # Selecciona                                             # noqa: E501

    # =========================================================
    # CRUD
    # =========================================================
    def on_nuevo(self):                                             # Nuevo alumno                                          # noqa: E501
        self.alumno_sel_id = None                                   # Limpia selección                                      # noqa: E501
        self.on_limpiar()                                           # Limpia form                                           # noqa: E501

    def on_limpiar(self):                                           # Limpia formulario                                     # noqa: E501
        self.var_tipo.set("Pregrado")                               # Default                                               # noqa: E501
        self.var_rut.set("")                                        # Limpia                                                # noqa: E501
        self.var_nombres.set("")                                    # Limpia                                                # noqa: E501
        self.var_apellidos.set("")                                  # Limpia                                                # noqa: E501
        self.var_email.set("")                                      # Limpia                                                # noqa: E501
        self.var_telefono.set("")                                   # Limpia                                                # noqa: E501
        self.var_periodo.set(f"{datetime.now().year}-1")            # Default                                               # noqa: E501
        self.var_activo.set(1)                                      # Activo                                                # noqa: E501

    def on_guardar(self):                                           # Guardar alumno (INSERT/UPDATE)                        # noqa: E501
        self._asegurar_schema_alumnos()                             # Asegura columnas antes de guardar                      # noqa: E501

        tipo = self.var_tipo.get().strip()                          # Tipo                                                  # noqa: E501
        rut = self.var_rut.get().strip()                            # RUT                                                   # noqa: E501
        nombres = self.var_nombres.get().strip()                    # Nombres                                               # noqa: E501
        apellidos = self.var_apellidos.get().strip()                # Apellidos                                             # noqa: E501
        email = self.var_email.get().strip()                        # Email                                                 # noqa: E501
        telefono = self.var_telefono.get().strip()                  # Teléfono                                              # noqa: E501
        periodo = self.var_periodo.get().strip()                    # Periodo                                               # noqa: E501
        estado = int(self.var_activo.get() or 0)                    # Estado                                                # noqa: E501

        uni = self._uni_sel()                                       # Universidad seleccionada                              # noqa: E501
        car = self._car_sel()                                       # Carrera seleccionada                                  # noqa: E501

        if not rut:                                                 # Validación                                            # noqa: E501
            messagebox.showwarning("Atención", "El RUT es obligatorio.")  # Aviso                                            # noqa: E501
            return                                                  # Sale                                                  # noqa: E501
        rut_norm = self._rut_norm(rut)                              # Normaliza                                             # noqa: E501
        if not rut_norm:                                            # Si inválido                                           # noqa: E501
            messagebox.showwarning("Atención", "RUT inválido. Ej: 14024474-0")  # Aviso                                        # noqa: E501
            return                                                  # Sale                                                  # noqa: E501
        if not nombres or not apellidos:                            # Validación                                            # noqa: E501
            messagebox.showwarning("Atención", "Nombres y Apellidos son obligatorios.")  # Aviso                                 # noqa: E501
            return                                                  # Sale                                                  # noqa: E501
        if not self._periodo_valido(periodo):                       # Periodo                                               # noqa: E501
            messagebox.showwarning("Atención", "Periodo inválido. Use YYYY-1 o YYYY-2 (ej: 2026-1).")  # Aviso               # noqa: E501
            return                                                  # Sale                                                  # noqa: E501
        if not uni or not car:                                      # Validación                                            # noqa: E501
            messagebox.showwarning("Atención", "Seleccione Universidad y Carrera.")  # Aviso                                  # noqa: E501
            return                                                  # Sale                                                  # noqa: E501

        nombre_busqueda = f"{apellidos} {nombres} {rut}".strip().lower()  # Texto búsqueda                                   # noqa: E501

        try:
            # Evita duplicado por rut_normalizado
            if self.alumno_sel_id is None:                          # Si es nuevo                                           # noqa: E501
                ex = self._q("SELECT alumno_id FROM alumnos WHERE rut_normalizado=?", (rut_norm,), fetch=True)  # Busca      # noqa: E501
                if ex:                                              # Si ya existe                                          # noqa: E501
                    messagebox.showwarning("Atención", "Ya existe un alumno con ese RUT (normalizado).")  # Aviso           # noqa: E501
                    return                                          # Sale                                                  # noqa: E501

                self._q(                                            # INSERT                                                 # noqa: E501
                    """
                    INSERT INTO alumnos(tipo_alumno, rut, rut_normalizado, nombres, apellidos, email, telefono,
                                       universidad_id, carrera_id, periodo, estado, nombre_busqueda)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (tipo, rut, rut_norm, nombres, apellidos, email, telefono, uni["id"], car["id"], periodo, estado, nombre_busqueda),
                )
                messagebox.showinfo("OK", "Alumno creado.")         # OK                                                     # noqa: E501
            else:                                                   # Si edita                                               # noqa: E501
                ex = self._q("SELECT alumno_id FROM alumnos WHERE rut_normalizado=? AND alumno_id<>?", (rut_norm, self.alumno_sel_id), fetch=True)  # Busca dup # noqa: E501
                if ex:                                              # Si duplicado                                           # noqa: E501
                    messagebox.showwarning("Atención", "Otro alumno ya usa ese RUT (normalizado).")     # Aviso             # noqa: E501
                    return                                          # Sale                                                  # noqa: E501

                self._q(                                            # UPDATE                                                 # noqa: E501
                    """
                    UPDATE alumnos
                    SET tipo_alumno=?, rut=?, rut_normalizado=?, nombres=?, apellidos=?, email=?, telefono=?,
                        universidad_id=?, carrera_id=?, periodo=?, estado=?, nombre_busqueda=?
                    WHERE alumno_id=?
                    """,
                    (tipo, rut, rut_norm, nombres, apellidos, email, telefono, uni["id"], car["id"], periodo, estado, nombre_busqueda, self.alumno_sel_id),
                )
                messagebox.showinfo("OK", "Alumno actualizado.")    # OK                                                     # noqa: E501

            self._cargar_listado(self.var_buscar.get().strip())     # Refresca listado                                       # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    def on_eliminar(self):                                          # Eliminar alumno                                       # noqa: E501
        if self.alumno_sel_id is None:                              # Debe seleccionar                                      # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un alumno para eliminar.")  # Aviso                         # noqa: E501
            return                                                  # Sale                                                  # noqa: E501
        if not messagebox.askyesno("Confirmar", "¿Eliminar alumno? (se eliminarán inscripciones y notas asociadas)"):  # Confirm      # noqa: E501
            return                                                  # Cancel                                                # noqa: E501
        try:
            self._q("DELETE FROM alumnos WHERE alumno_id=?", (self.alumno_sel_id,))  # Delete                          # noqa: E501
            self.alumno_sel_id = None                               # Limpia selección                                      # noqa: E501
            self.on_limpiar()                                       # Limpia form                                           # noqa: E501
            self._cargar_listado(self.var_buscar.get().strip())     # Refresca listado                                      # noqa: E501
            messagebox.showinfo("OK", "Alumno eliminado.")          # OK                                                     # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    # =========================================================
    # Crear universidad/carrera rápido
    # =========================================================
    def on_nueva_universidad(self):                                 # Crea universidad desde UI                              # noqa: E501
        nombre = simpledialog.askstring("Nueva universidad", "Nombre de la universidad:")  # Pide nombre                # noqa: E501
        if not nombre:                                              # Cancel                                                 # noqa: E501
            return                                                  # Sale                                                   # noqa: E501
        nombre = nombre.strip()                                     # Limpia                                                 # noqa: E501
        if not nombre:                                              # Vacío                                                  # noqa: E501
            return                                                  # Sale                                                   # noqa: E501
        try:
            self._q("INSERT INTO universidades(nombre) VALUES(?)", (nombre,))  # Inserta                              # noqa: E501
            self._cargar_universidades()                            # Refresca combos                                       # noqa: E501
            self.var_uni.set(nombre)                                # Selecciona nueva                                      # noqa: E501
            self.on_uni_change()                                    # Carga carreras                                         # noqa: E501
            messagebox.showinfo("OK", "Universidad creada.")        # OK                                                     # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    def on_nueva_carrera(self):                                     # Crea carrera desde UI                                  # noqa: E501
        uni = self._uni_sel()                                       # Uni seleccionada                                       # noqa: E501
        if not uni:                                                 # Validación                                             # noqa: E501
            messagebox.showwarning("Atención", "Seleccione una universidad antes de crear una carrera.")  # Aviso          # noqa: E501
            return                                                  # Sale                                                   # noqa: E501
        nombre = simpledialog.askstring("Nueva carrera", "Nombre de la carrera:")  # Pide nombre                      # noqa: E501
        if not nombre:                                              # Cancel                                                 # noqa: E501
            return                                                  # Sale                                                   # noqa: E501
        nombre = nombre.strip()                                     # Limpia                                                 # noqa: E501
        if not nombre:                                              # Vacío                                                  # noqa: E501
            return                                                  # Sale                                                   # noqa: E501
        try:
            self._q("INSERT INTO carreras(universidad_id, nombre) VALUES(?,?)", (uni["id"], nombre))  # Inserta            # noqa: E501
            self._cargar_carreras(uni["id"])                        # Refresca carreras                                     # noqa: E501
            self.var_car.set(nombre)                                # Selecciona nueva                                      # noqa: E501
            messagebox.showinfo("OK", "Carrera creada.")            # OK                                                     # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    # =========================================================
    # Exportaciones
    # =========================================================
    def _rows_export(self):                                         # Obtiene filas exportables                              # noqa: E501
        t = self.var_buscar.get().strip()                           # Filtro                                                 # noqa: E501
        if t:                                                       # Con filtro                                             # noqa: E501
            q = f"%{t}%"                                           # Like                                                   # noqa: E501
            return self._q(
                """
                SELECT
                  a.tipo_alumno, a.rut, a.nombres, a.apellidos, COALESCE(a.email,'') AS email, COALESCE(a.telefono,'') AS telefono,
                  COALESCE(u.nombre,'') AS universidad, COALESCE(ca.nombre,'') AS carrera, COALESCE(a.periodo,'') AS periodo,
                  COALESCE(a.estado,1) AS estado
                FROM alumnos a
                LEFT JOIN universidades u ON u.universidad_id=a.universidad_id
                LEFT JOIN carreras ca ON ca.carrera_id=a.carrera_id
                WHERE
                  COALESCE(a.rut,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.nombres,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.apellidos,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.email,'') LIKE ? COLLATE NOCASE OR
                  COALESCE(a.nombre_busqueda,'') LIKE ? COLLATE NOCASE
                ORDER BY a.apellidos, a.nombres
                """,
                (q, q, q, q, q),
                fetch=True,
            ) or []
        else:                                                       # Sin filtro                                             # noqa: E501
            return self._q(
                """
                SELECT
                  a.tipo_alumno, a.rut, a.nombres, a.apellidos, COALESCE(a.email,'') AS email, COALESCE(a.telefono,'') AS telefono,
                  COALESCE(u.nombre,'') AS universidad, COALESCE(ca.nombre,'') AS carrera, COALESCE(a.periodo,'') AS periodo,
                  COALESCE(a.estado,1) AS estado
                FROM alumnos a
                LEFT JOIN universidades u ON u.universidad_id=a.universidad_id
                LEFT JOIN carreras ca ON ca.carrera_id=a.carrera_id
                ORDER BY a.apellidos, a.nombres
                """,
                fetch=True,
            ) or []

    def on_exportar_excel(self):                                    # Exporta alumnos a Excel                               # noqa: E501
        try:
            rows = self._rows_export()                              # Datos                                                  # noqa: E501
            ruta = filedialog.asksaveasfilename(                    # Ruta                                                   # noqa: E501
                title="Guardar Excel (Alumnos)",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                            # Cancel                                                 # noqa: E501
                return                                              # Sale                                                   # noqa: E501

            wb = Workbook()                                         # Libro                                                  # noqa: E501
            ws = wb.active                                          # Hoja                                                   # noqa: E501
            ws.title = "Alumnos"                                    # Nombre                                                 # noqa: E501

            ws.append(["TipoAlumno", "RUT", "Nombres", "Apellidos", "Email", "Telefono", "Universidad", "Carrera", "Periodo", "Estado"])  # Header # noqa: E501
            for r in rows:                                          # Filas                                                  # noqa: E501
                ws.append([r["tipo_alumno"], r["rut"], r["nombres"], r["apellidos"], r["email"], r["telefono"], r["universidad"], r["carrera"], r["periodo"], int(r["estado"] or 1)])  # Row # noqa: E501

            wb.save(ruta)                                           # Guarda                                                 # noqa: E501
            messagebox.showinfo("OK", "Excel exportado correctamente.")  # OK                                                  # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    def on_exportar_pdf(self):                                      # Exporta alumnos a PDF                                 # noqa: E501
        try:
            rows = self._rows_export()                              # Datos                                                  # noqa: E501
            ruta = filedialog.asksaveasfilename(                    # Ruta                                                   # noqa: E501
                title="Guardar PDF (Alumnos)",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                            # Cancel                                                 # noqa: E501
                return                                              # Sale                                                   # noqa: E501

            c = canvas.Canvas(ruta, pagesize=A4)                    # Canvas                                                 # noqa: E501
            w, h = A4                                               # Dimensiones                                            # noqa: E501
            y = h - 60                                              # Cursor Y                                               # noqa: E501

            c.setFont("Helvetica-Bold", 13)                         # Fuente título                                          # noqa: E501
            c.drawString(40, y, "Listado de Alumnos")               # Título                                                 # noqa: E501
            y -= 20                                                 # Baja                                                   # noqa: E501

            c.setFont("Helvetica-Bold", 9)                          # Cabecera                                               # noqa: E501
            c.drawString(40, y, "RUT")                              # Col                                                    # noqa: E501
            c.drawString(120, y, "Alumno")                          # Col                                                    # noqa: E501
            c.drawString(300, y, "Periodo")                         # Col                                                    # noqa: E501
            c.drawString(360, y, "Tipo")                            # Col                                                    # noqa: E501
            y -= 12                                                 # Baja                                                   # noqa: E501

            c.setFont("Helvetica", 9)                               # Normal                                                 # noqa: E501
            for r in rows:                                          # Filas                                                  # noqa: E501
                if y < 60:                                          # Salto página                                           # noqa: E501
                    c.showPage()                                    # Nueva página                                           # noqa: E501
                    y = h - 60                                      # Reinicia                                               # noqa: E501
                    c.setFont("Helvetica", 9)                       # Fuente                                                 # noqa: E501
                alumno = f"{r['apellidos']} {r['nombres']}".strip() # Nombre completo                                        # noqa: E501
                c.drawString(40, y, str(r["rut"])[:12])             # Rut                                                    # noqa: E501
                c.drawString(120, y, alumno[:30])                   # Alumno                                                 # noqa: E501
                c.drawString(300, y, str(r["periodo"])[:10])        # Periodo                                                # noqa: E501
                c.drawString(360, y, str(r["tipo_alumno"])[:12])    # Tipo                                                   # noqa: E501
                y -= 12                                             # Baja                                                   # noqa: E501

            c.save()                                                # Guarda PDF                                             # noqa: E501
            messagebox.showinfo("OK", "PDF exportado correctamente.")  # OK                                                  # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                   # Error                                                  # noqa: E501

    # =========================================================
    # Importar Excel (crea uni/carrera si no existen)
    # =========================================================
    def on_importar_excel(self):                                    # Importa alumnos desde Excel                           # noqa: E501
        self._asegurar_schema_alumnos()                             # Asegura columnas                                      # noqa: E501

        ruta = filedialog.askopenfilename(                          # Selecciona archivo                                    # noqa: E501
            title="Importar Excel (Alumnos)",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=".",
        )
        if not ruta:                                                # Cancel                                                 # noqa: E501
            return                                                  # Sale                                                   # noqa: E501

        try:
            wb = load_workbook(ruta)                                # Abre libro                                             # noqa: E501
            ws = wb.active                                          # Hoja activa                                            # noqa: E501

            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Cabecera                                       # noqa: E501
            headers = [str(h or "").strip().casefold() for h in header]           # Normaliza                                      # noqa: E501

            def hidx(*names: str) -> int:                           # Busca índice por alias                                # noqa: E501
                for n in names:                                     # Recorre                                                # noqa: E501
                    k = n.casefold()                                # Normaliza                                              # noqa: E501
                    if k in headers:                                # Si existe                                              # noqa: E501
                        return headers.index(k)                     # Retorna índice                                         # noqa: E501
                return -1                                           # No encontrado                                          # noqa: E501

            i_tipo = hidx("tipoalumno", "tipo_alumno", "tipo")      # Tipo                                                   # noqa: E501
            i_rut = hidx("rut")                                     # RUT                                                    # noqa: E501
            i_nom = hidx("nombres", "nombre")                       # Nombres                                                # noqa: E501
            i_ape = hidx("apellidos", "apellido")                   # Apellidos                                              # noqa: E501
            i_email = hidx("email", "correo")                       # Email                                                  # noqa: E501
            i_tel = hidx("telefono", "teléfono")                    # Teléfono                                               # noqa: E501
            i_uni = hidx("universidad")                             # Universidad                                            # noqa: E501
            i_car = hidx("carrera")                                 # Carrera                                                # noqa: E501
            i_per = hidx("periodo", "período")                      # Periodo                                                # noqa: E501
            i_est = hidx("estado")                                  # Estado (0/1)                                           # noqa: E501

            if i_rut == -1 or i_nom == -1 or i_ape == -1 or i_uni == -1 or i_car == -1 or i_per == -1:  # Validación # noqa: E501
                raise ValueError("Excel debe incluir: RUT, Nombres, Apellidos, Universidad, Carrera, Periodo. (Opcional: TipoAlumno, Email, Telefono, Estado).")  # Error # noqa: E501

            conn = obtener_conexion()                               # Conexión única para performance                       # noqa: E501
            try:
                cur = conn.cursor()                                 # Cursor                                                 # noqa: E501

                total = 0                                           # Total                                                  # noqa: E501
                ok = 0                                              # OK                                                     # noqa: E501
                errores = []                                        # Errores                                                # noqa: E501

                for fila_n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Itera datos        # noqa: E501
                    total += 1                                      # Cuenta                                                 # noqa: E501
                    try:
                        tipo = str(row[i_tipo] or "Pregrado").strip() if i_tipo != -1 else "Pregrado"  # Tipo          # noqa: E501
                        rut = str(row[i_rut] or "").strip()         # RUT                                                    # noqa: E501
                        nombres = str(row[i_nom] or "").strip()     # Nombres                                                # noqa: E501
                        apellidos = str(row[i_ape] or "").strip()   # Apellidos                                              # noqa: E501
                        email = str(row[i_email] or "").strip() if i_email != -1 else ""  # Email                    # noqa: E501
                        tel = str(row[i_tel] or "").strip() if i_tel != -1 else ""        # Teléfono                 # noqa: E501
                        uni = str(row[i_uni] or "").strip()         # Universidad                                            # noqa: E501
                        car = str(row[i_car] or "").strip()         # Carrera                                                # noqa: E501
                        per = str(row[i_per] or "").strip()         # Periodo                                                # noqa: E501
                        est = int(row[i_est]) if (i_est != -1 and row[i_est] is not None) else 1  # Estado            # noqa: E501

                        if not rut or not nombres or not apellidos or not uni or not car or not per:  # Validación      # noqa: E501
                            raise ValueError("Faltan datos obligatorios.")                            # Error           # noqa: E501
                        if not self._periodo_valido(per):         # Periodo válido                                        # noqa: E501
                            raise ValueError("Periodo inválido (use YYYY-1 o YYYY-2).")               # Error           # noqa: E501

                        rut_norm = self._rut_norm(rut)             # Normaliza                                              # noqa: E501
                        if not rut_norm:                           # Si inválido                                            # noqa: E501
                            raise ValueError("RUT inválido.")      # Error                                                  # noqa: E501

                        # Universidad: buscar/crear
                        cur.execute("SELECT universidad_id FROM universidades WHERE nombre=?", (uni,))  # Busca uni      # noqa: E501
                        r = cur.fetchone()                                                                   # Fila         # noqa: E501
                        if r:                                                                                # Existe       # noqa: E501
                            uni_id = int(r[0])                                                               # ID           # noqa: E501
                        else:                                                                                # No existe    # noqa: E501
                            cur.execute("INSERT INTO universidades(nombre) VALUES(?)", (uni,))               # Inserta      # noqa: E501
                            uni_id = int(cur.lastrowid)                                                      # ID nuevo     # noqa: E501

                        # Carrera: buscar/crear (por uni)
                        cur.execute("SELECT carrera_id FROM carreras WHERE universidad_id=? AND nombre=?", (uni_id, car))  # Busca # noqa: E501
                        r = cur.fetchone()                                                                   # Fila         # noqa: E501
                        if r:                                                                                # Existe       # noqa: E501
                            car_id = int(r[0])                                                               # ID           # noqa: E501
                        else:                                                                                # No existe    # noqa: E501
                            cur.execute("INSERT INTO carreras(universidad_id, nombre) VALUES(?,?)", (uni_id, car))  # Ins # noqa: E501
                            car_id = int(cur.lastrowid)                                                      # ID nuevo     # noqa: E501

                        nombre_busqueda = f"{apellidos} {nombres} {rut}".strip().lower()                    # Busqueda      # noqa: E501

                        # UPSERT por rut_normalizado (si existe, actualiza)
                        cur.execute("SELECT alumno_id FROM alumnos WHERE rut_normalizado=?", (rut_norm,))   # Busca alumno  # noqa: E501
                        ex = cur.fetchone()                                                                  # Resultado     # noqa: E501
                        if ex:                                                                              # Si existe     # noqa: E501
                            alumno_id = int(ex[0])                                                          # ID            # noqa: E501
                            cur.execute(
                                """
                                UPDATE alumnos
                                SET tipo_alumno=?, rut=?, nombres=?, apellidos=?, email=?, telefono=?,
                                    universidad_id=?, carrera_id=?, periodo=?, estado=?, nombre_busqueda=?, rut_normalizado=?
                                WHERE alumno_id=?
                                """,
                                (tipo, rut, nombres, apellidos, email, tel, uni_id, car_id, per, est, nombre_busqueda, rut_norm, alumno_id),
                            )                                                                               # Update         # noqa: E501
                        else:                                                                              # Si no existe   # noqa: E501
                            cur.execute(
                                """
                                INSERT INTO alumnos(tipo_alumno, rut, rut_normalizado, nombres, apellidos, email, telefono,
                                                   universidad_id, carrera_id, periodo, estado, nombre_busqueda)
                                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                                """,
                                (tipo, rut, rut_norm, nombres, apellidos, email, tel, uni_id, car_id, per, est, nombre_busqueda),
                            )                                                                               # Insert         # noqa: E501

                        ok += 1                                                                            # OK fila        # noqa: E501
                    except Exception as e:
                        errores.append(f"Fila {fila_n}: {e}")                                               # Guarda error   # noqa: E501

                conn.commit()                                                                              # Commit         # noqa: E501
            finally:
                conn.close()                                                                               # Close          # noqa: E501

            self._cargar_universidades()                                                                    # Refresca combos # noqa: E501
            self._cargar_listado(self.var_buscar.get().strip())                                             # Refresca listado # noqa: E501

            msg = f"Importación finalizada.\nProcesadas: {total}\nOK: {ok}\nErrores: {len(errores)}"       # Resumen        # noqa: E501
            if errores:                                                                                    # Si errores     # noqa: E501
                msg += "\n\nPrimeros errores:\n" + "\n".join(errores[:10])                                  # Muestra 10     # noqa: E501
            messagebox.showinfo("Resultado", msg)                                                           # Muestra        # noqa: E501

        except Exception as e:
            messagebox.showerror("Error", str(e))                                                           # Error general  # noqa: E501
