# ============================================  # Separador visual
# modulos/ui_notas.py                            # Página "Notas" (UI única, responsiva)
# - Curso arriba + suma de porcentajes            # Barra superior compacta
# - Izquierda: alumnos inscritos                  # Treeview
# - Derecha: tabla única evaluaciones+notas       # Treeview 3 columnas iguales (1/3, 1/3, 1/3)
# - Doble click edita Nota                        # Editor in-place
# - Panel de casillas para editar nota            # Entry + Guardar
# - Exportar Excel / PDF                          # Reporte completo por curso
# - Importar Excel                                # Carga masiva de notas (largo o matriz)
# ============================================  # Separador visual

import tkinter as tk                                                     # Tkinter base                                                     # noqa: E501
from tkinter import ttk, messagebox, filedialog, simpledialog            # Widgets + diálogos                                               # noqa: E501

from openpyxl import Workbook                                            # Excel export                                                      # noqa: E501
from openpyxl import load_workbook                                       # Excel import                                                      # noqa: E501

from reportlab.pdfgen import canvas                                      # PDF                                                               # noqa: E501
from reportlab.lib.pagesizes import A4                                   # Tamaño página                                                     # noqa: E501

from .bd_sqlite import obtener_conexion                                  # Conexión SQLite                                                   # noqa: E501
from .config import carpeta_exports                                      # Carpeta sugerida para exportaciones                               # noqa: E501


class PaginaNotas(ttk.Frame):                                            # Página Notas integrada a UI única                                 # noqa: E501
    def __init__(self, master):                                          # Constructor                                                       # noqa: E501
        super().__init__(master)                                         # Inicializa Frame                                                  # noqa: E501

        self.curso_sel_id = None                                         # ID curso seleccionado                                             # noqa: E501
        self.inscripcion_sel_id = None                                   # ID inscripción seleccionada                                       # noqa: E501

        self._cursos = []                                                # Lista de cursos (dicts)                                           # noqa: E501
        self._cursos_display = []                                        # Strings para combobox                                             # noqa: E501

        self._editor_entry = None                                        # Entry flotante para doble click                                  # noqa: E501
        self._editor_item = None                                         # Item actual editándose                                            # noqa: E501

        self.var_eval_nombre = tk.StringVar()                            # Nombre evaluación en panel                                        # noqa: E501
        self.var_eval_porcentaje = tk.StringVar()                        # Porcentaje en panel                                               # noqa: E501
        self.var_nota = tk.StringVar()                                   # Nota en panel                                                     # noqa: E501

        self._crear_ui()                                                 # Construye UI                                                      # noqa: E501
        self._cargar_cursos()                                            # Carga cursos                                                      # noqa: E501

    def on_show(self):                                                   # Hook cuando la página se muestra                                  # noqa: E501
        self._cargar_cursos()                                            # Refresca cursos (por si se crearon nuevos)                        # noqa: E501

    # =========================================================
    # Statusbar (si ventana principal tiene set_status)
    # =========================================================
    def _status(self, texto: str):                                       # Envía texto a barra de estado                                     # noqa: E501
        top = self.winfo_toplevel()                                      # Obtiene ventana principal                                         # noqa: E501
        if hasattr(top, "set_status") and callable(getattr(top, "set_status")):  # Si existe set_status                      # noqa: E501
            top.set_status(texto)                                        # Actualiza estado                                                  # noqa: E501

    # =========================================================
    # Helper SQL
    # =========================================================
    def _q(self, sql: str, params: tuple = (), fetch: bool = False):     # Ejecuta SQL con parámetros                                         # noqa: E501
        conn = obtener_conexion()                                        # Abre conexión                                                     # noqa: E501
        try:                                                             # Bloque seguro                                                     # noqa: E501
            cur = conn.cursor()                                          # Cursor                                                            # noqa: E501
            cur.execute(sql, params)                                     # Ejecuta                                                           # noqa: E501
            if fetch:                                                    # Si pide resultados                                                # noqa: E501
                return cur.fetchall()                                    # Devuelve filas                                                    # noqa: E501
            conn.commit()                                                # Confirma cambios                                                  # noqa: E501
            return None                                                  # Devuelve None                                                     # noqa: E501
        finally:                                                         # Siempre                                                           # noqa: E501
            conn.close()                                                 # Cierra conexión                                                   # noqa: E501

    # =========================================================
    # UI
    # =========================================================
    def _crear_ui(self):                                                 # Construye la UI                                                   # noqa: E501
        self.columnconfigure(0, weight=1)                                # Crece en ancho                                                    # noqa: E501
        self.rowconfigure(2, weight=1)                                   # ✅ Solo contenido principal crece en alto                          # noqa: E501

        ttk.Label(self, text="Notas", font=("Segoe UI", 16, "bold")).grid(  # Título                                                          # noqa: E501
            row=0, column=0, sticky="w", padx=12, pady=(12, 6)           # Posición                                                          # noqa: E501
        )

        # ---------------- Barra superior compacta ----------------
        top = ttk.Frame(self)                                            # Frame superior                                                    # noqa: E501
        top.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))      # No se estira en alto                                              # noqa: E501
        top.columnconfigure(1, weight=1)                                 # Combo curso se expande en ancho                                   # noqa: E501

        ttk.Label(top, text="Curso:").grid(row=0, column=0, sticky="w")  # Label curso                                                       # noqa: E501

        self.var_curso = tk.StringVar()                                  # Variable de curso                                                 # noqa: E501
        self.cmb_curso = ttk.Combobox(top, textvariable=self.var_curso, state="readonly")  # Combo curso                                # noqa: E501
        self.cmb_curso.grid(row=0, column=1, sticky="ew", padx=(8, 10))  # Ubicación                                                         # noqa: E501
        self.cmb_curso.bind("<<ComboboxSelected>>", lambda e: self.on_curso_change())  # Evento cambio                            # noqa: E501

        self.lbl_suma = ttk.Label(top, text="Suma %: 0.00")              # Suma porcentajes                                                  # noqa: E501
        self.lbl_suma.grid(row=0, column=2, sticky="w", padx=(0, 10))    # Ubicación                                                         # noqa: E501

        ttk.Button(top, text="Inscribir", command=self.on_inscribir).grid(row=0, column=3, padx=4)       # Inscribir                # noqa: E501
        ttk.Button(top, text="Quitar", command=self.on_desinscribir).grid(row=0, column=4, padx=4)       # Quitar                   # noqa: E501

        ttk.Separator(top, orient="vertical").grid(row=0, column=5, sticky="ns", padx=10)                # Separador                # noqa: E501

        ttk.Button(top, text="Exportar Excel", command=self.on_exportar_excel).grid(row=0, column=6, padx=4)  # Excel                 # noqa: E501
        ttk.Button(top, text="Exportar PDF", command=self.on_exportar_pdf).grid(row=0, column=7, padx=4)      # PDF                   # noqa: E501
        ttk.Button(top, text="Importar Excel", command=self.on_importar_excel).grid(row=0, column=8, padx=4)  # Import                # noqa: E501

        # ---------------- Contenido principal: PanedWindow ----------------
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)         # Divisor arrastrable                                               # noqa: E501
        self.paned.grid(row=2, column=0, sticky="nsew", padx=12, pady=12)  # Ocupa todo                                                       # noqa: E501

        # Panel izquierdo (inscritos)
        self.panel_izq = ttk.Frame(self.paned)                           # Panel izq                                                         # noqa: E501
        self.panel_izq.columnconfigure(0, weight=1)                      # Tabla crece                                                       # noqa: E501
        self.panel_izq.rowconfigure(0, weight=1)                         # Tabla crece                                                       # noqa: E501

        # Panel derecho (notas + formulario)
        self.panel_der = ttk.Frame(self.paned)                           # Panel der                                                         # noqa: E501
        self.panel_der.columnconfigure(0, weight=1)                      # Contenido crece                                                   # noqa: E501
        self.panel_der.rowconfigure(1, weight=1)                         # ✅ La tabla de notas crece                                          # noqa: E501

        self.paned.add(self.panel_izq, weight=2)                          # Izq                                                              # noqa: E501
        self.paned.add(self.panel_der, weight=3)                          # Der                                                              # noqa: E501

        try:                                                              # Mínimos panes                                                     # noqa: E501
            self.paned.paneconfigure(self.panel_izq, minsize=420)         # Min izq                                                           # noqa: E501
            self.paned.paneconfigure(self.panel_der, minsize=620)         # Min der                                                           # noqa: E501
        except Exception:                                                 # Si no soporta                                                     # noqa: E501
            pass                                                          # No rompe                                                          # noqa: E501

        # ---------------- IZQUIERDA: inscritos ----------------
        lf_izq = ttk.LabelFrame(self.panel_izq, text="Alumnos inscritos")  # Marco                                                            # noqa: E501
        lf_izq.grid(row=0, column=0, sticky="nsew", padx=(0, 10))          # Ubicación                                                        # noqa: E501
        lf_izq.columnconfigure(0, weight=1)                                # Tabla crece                                                      # noqa: E501
        lf_izq.rowconfigure(0, weight=1)                                   # Tabla crece                                                      # noqa: E501

        self.tree_insc = ttk.Treeview(                                     # Tree inscritos                                                   # noqa: E501
            lf_izq,
            columns=("insc_id", "rut", "alumno", "prom"),
            show="headings",
            selectmode="browse",
        )
        self.tree_insc.heading("insc_id", text="InscID")                   # Encabezado                                                       # noqa: E501
        self.tree_insc.heading("rut", text="RUT")                          # Encabezado                                                       # noqa: E501
        self.tree_insc.heading("alumno", text="Alumno")                    # Encabezado                                                       # noqa: E501
        self.tree_insc.heading("prom", text="Prom.")                       # Encabezado                                                       # noqa: E501

        self.tree_insc.column("insc_id", width=70, anchor="center")        # Col                                                             # noqa: E501
        self.tree_insc.column("rut", width=120, anchor="w")                # Col                                                             # noqa: E501
        self.tree_insc.column("alumno", width=230, anchor="w")             # Col                                                             # noqa: E501
        self.tree_insc.column("prom", width=80, anchor="center")           # Col                                                             # noqa: E501

        sb1 = ttk.Scrollbar(lf_izq, orient="vertical", command=self.tree_insc.yview)  # Scroll                                            # noqa: E501
        self.tree_insc.configure(yscrollcommand=sb1.set)                   # Enlaza scroll                                                    # noqa: E501

        self.tree_insc.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)  # Ubicación                                                      # noqa: E501
        sb1.grid(row=0, column=1, sticky="ns", pady=8)                     # Ubicación                                                        # noqa: E501

        self.tree_insc.bind("<<TreeviewSelect>>", lambda e: self.on_select_inscrito())  # Selección -> notas                            # noqa: E501

        # ---------------- DERECHA: notas del alumno ----------------
        self.lbl_alumno = ttk.Label(self.panel_der, text="Seleccione un alumno inscrito.", font=("Segoe UI", 11, "bold"))  # Header                             # noqa: E501
        self.lbl_alumno.grid(row=0, column=0, sticky="w", padx=10, pady=(0, 10))                                           # Ubicación                           # noqa: E501

        # Tabla única: Evaluación | % | Nota  (3 columnas visibles iguales)
        tabla_frame = ttk.Frame(self.panel_der)                            # Frame tabla + scrollbar                                          # noqa: E501
        tabla_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))  # Ubicación                                                    # noqa: E501
        tabla_frame.columnconfigure(0, weight=1)                            # Tabla se expande                                                  # noqa: E501
        tabla_frame.rowconfigure(0, weight=1)                               # Tabla se expande                                                  # noqa: E501

        self.tree_notas = ttk.Treeview(                                    # Tree notas                                                       # noqa: E501
            tabla_frame,
            columns=("eval_id", "evaluacion", "porcentaje", "nota"),
            show="headings",
            selectmode="browse",
        )
        self.tree_notas.heading("eval_id", text="EvalID")                  # Encabezado interno                                               # noqa: E501
        self.tree_notas.heading("evaluacion", text="Evaluación")           # Encabezado                                                       # noqa: E501
        self.tree_notas.heading("porcentaje", text="%")                    # Encabezado                                                       # noqa: E501
        self.tree_notas.heading("nota", text="Nota")                       # Encabezado                                                       # noqa: E501

        self.tree_notas.column("eval_id", width=0, stretch=False)          # Oculta EvalID                                                    # noqa: E501
        self.tree_notas.column("evaluacion", width=200, anchor="w", stretch=True)   # Inicial (se ajusta luego)                              # noqa: E501
        self.tree_notas.column("porcentaje", width=200, anchor="center", stretch=True)  # Inicial (se ajusta luego)                          # noqa: E501
        self.tree_notas.column("nota", width=200, anchor="center", stretch=True)   # Inicial (se ajusta luego)                                  # noqa: E501

        sb2 = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tree_notas.yview)  # Scroll                                        # noqa: E501
        self.tree_notas.configure(yscrollcommand=sb2.set)                 # Enlaza scroll                                                    # noqa: E501

        self.tree_notas.grid(row=0, column=0, sticky="nsew")              # Tabla ocupa espacio                                              # noqa: E501
        sb2.grid(row=0, column=1, sticky="ns")                            # Scroll a la derecha                                              # noqa: E501

        self.tree_notas.bind("<Double-1>", self.on_doble_click_nota)      # Doble click edita nota                                           # noqa: E501
        self.tree_notas.bind("<<TreeviewSelect>>", lambda e: self.on_select_evaluacion())  # Selección -> panel                    # noqa: E501

        # ✅ Ajuste EXACTO de 1/3 por columna visible (Evaluación / % / Nota)
        self.tree_notas.bind("<Configure>", lambda e: self._ajustar_columnas_tercios())  # Cada resize ajusta anchos                   # noqa: E501

        # Panel casillas
        panel = ttk.LabelFrame(self.panel_der, text="Ingreso de nota (casillas)")  # Marco                                               # noqa: E501
        panel.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))             # Ubicación                                          # noqa: E501
        panel.columnconfigure(1, weight=1)                                          # Columna 1 crece                                   # noqa: E501

        ttk.Label(panel, text="Evaluación:").grid(row=0, column=0, sticky="w", padx=8, pady=6)  # Label                                         # noqa: E501
        ttk.Entry(panel, textvariable=self.var_eval_nombre, state="readonly").grid(row=0, column=1, sticky="ew", padx=8, pady=6)  # Entry  # noqa: E501

        ttk.Label(panel, text="%:").grid(row=0, column=2, sticky="w", padx=8, pady=6)           # Label                                         # noqa: E501
        ttk.Entry(panel, textvariable=self.var_eval_porcentaje, width=10, state="readonly").grid(row=0, column=3, sticky="w", padx=8, pady=6)  # Entry # noqa: E501

        ttk.Label(panel, text="Nota:").grid(row=1, column=0, sticky="w", padx=8, pady=6)        # Label                                         # noqa: E501
        ttk.Entry(panel, textvariable=self.var_nota, width=15).grid(row=1, column=1, sticky="w", padx=8, pady=6)  # Entry                                    # noqa: E501

        btn_panel = ttk.Frame(panel)                                          # Frame botones panel                                               # noqa: E501
        btn_panel.grid(row=1, column=2, columnspan=2, sticky="e", padx=8, pady=6)  # Ubicación                                                     # noqa: E501
        ttk.Button(btn_panel, text="Guardar nota", command=self.on_guardar_nota_panel).pack(side="left", padx=5)  # Guardar              # noqa: E501
        ttk.Button(btn_panel, text="Limpiar", command=self.on_limpiar_panel).pack(side="left", padx=5)           # Limpiar              # noqa: E501

        # Botones generales
        acciones = ttk.Frame(self.panel_der)                                  # Frame acciones                                                  # noqa: E501
        acciones.grid(row=3, column=0, sticky="w", padx=10, pady=6)           # Ubicación                                                       # noqa: E501
        ttk.Button(acciones, text="Guardar todo (tabla)", command=self.on_guardar_todo).pack(side="left", padx=5)  # Guardar todo          # noqa: E501
        ttk.Button(acciones, text="Recalcular promedio", command=self.on_recalcular).pack(side="left", padx=5)     # Recalcular            # noqa: E501

        self.lbl_prom = ttk.Label(self.panel_der, text="Promedio ponderado: -", font=("Segoe UI", 11))  # Label promedio                         # noqa: E501
        self.lbl_prom.grid(row=4, column=0, sticky="w", padx=10, pady=(0, 10))                           # Ubicación                              # noqa: E501

    def _ajustar_columnas_tercios(self):                                      # Ajusta 3 columnas visibles en tercios exactos                     # noqa: E501
        total = self.tree_notas.winfo_width()                                 # Ancho actual del Treeview                                         # noqa: E501
        if total <= 50:                                                       # Si aún no está listo (UI inicial)                                 # noqa: E501
            return                                                            # Sale                                                              # noqa: E501
        # Restamos un margen aproximado (borde/scroll interno) para evitar corte
        usable = max(300, total - 20)                                         # Ancho usable mínimo                                               # noqa: E501
        tercio = max(90, usable // 3)                                         # Cada columna 1/3                                                  # noqa: E501
        self.tree_notas.column("evaluacion", width=tercio)                    # 1/3 evaluación                                                    # noqa: E501
        self.tree_notas.column("porcentaje", width=tercio)                    # 1/3 porcentaje                                                    # noqa: E501
        self.tree_notas.column("nota", width=tercio)                          # 1/3 nota                                                          # noqa: E501

    # =========================================================
    # Cursos
    # =========================================================
    def _cargar_cursos(self):                                                 # Carga cursos detallados en combobox                               # noqa: E501
        rows = self._q(                                                       # Query cursos                                                      # noqa: E501
            """
            SELECT
              c.curso_id,
              c.nombre AS curso_nombre,
              COALESCE(c.codigo,'') AS codigo,
              c.periodo,
              ca.nombre AS carrera_nombre,
              u.nombre AS universidad_nombre
            FROM cursos c
            JOIN carreras ca ON ca.carrera_id=c.carrera_id
            JOIN universidades u ON u.universidad_id=ca.universidad_id
            ORDER BY u.nombre, ca.nombre, c.periodo DESC, c.nombre
            """,
            fetch=True,
        ) or []

        self._cursos = []                                                     # Reinicia lista                                                    # noqa: E501
        self._cursos_display = []                                             # Reinicia display                                                  # noqa: E501

        for r in rows:                                                        # Recorre cursos                                                     # noqa: E501
            item = {                                                          # Dict curso                                                         # noqa: E501
                "curso_id": int(r["curso_id"]),                               # ID                                                                  # noqa: E501
                "universidad_nombre": r["universidad_nombre"],                # Universidad                                                        # noqa: E501
                "carrera_nombre": r["carrera_nombre"],                        # Carrera                                                            # noqa: E501
                "periodo": r["periodo"],                                      # Periodo                                                            # noqa: E501
                "curso_nombre": r["curso_nombre"],                            # Curso                                                              # noqa: E501
                "codigo": r["codigo"],                                        # Código                                                             # noqa: E501
            }
            self._cursos.append(item)                                         # Guarda                                                             # noqa: E501

            txt = f"{item['universidad_nombre']} | {item['carrera_nombre']} | {item['periodo']} | {item['curso_nombre']}"  # Texto combobox # noqa: E501
            if item["codigo"]:                                                # Si hay código                                                      # noqa: E501
                txt += f" ({item['codigo']})"                                 # Agrega                                                             # noqa: E501
            self._cursos_display.append(txt)                                  # Guarda                                                             # noqa: E501

        self.cmb_curso["values"] = self._cursos_display                       # Set values                                                         # noqa: E501

        if self._cursos_display:                                              # Si hay cursos                                                      # noqa: E501
            if self.var_curso.get().strip() not in self._cursos_display:      # Si selección inválida                                              # noqa: E501
                self.var_curso.set(self._cursos_display[0])                   # Selecciona primero                                                 # noqa: E501
            self.on_curso_change()                                            # Carga data                                                         # noqa: E501
        else:                                                                 # Si no hay cursos                                                   # noqa: E501
            self.var_curso.set("")                                            # Limpia                                                             # noqa: E501
            self.curso_sel_id = None                                          # Sin curso                                                          # noqa: E501
            self._limpiar_todo()                                              # Limpia UI                                                          # noqa: E501

    def _curso_info_sel(self):                                                # Retorna dict del curso seleccionado                               # noqa: E501
        txt = self.var_curso.get().strip()                                    # Texto seleccionado                                                 # noqa: E501
        if not txt or txt not in self._cursos_display:                        # Si no existe                                                       # noqa: E501
            return None                                                       # None                                                               # noqa: E501
        idx = self._cursos_display.index(txt)                                 # Índice                                                             # noqa: E501
        return self._cursos[idx]                                              # Dict curso                                                         # noqa: E501

    def on_curso_change(self):                                                # Al cambiar curso                                                   # noqa: E501
        info = self._curso_info_sel()                                         # Obtiene dict                                                       # noqa: E501
        if not info:                                                          # Si no hay                                                          # noqa: E501
            self.curso_sel_id = None                                          # Limpia                                                             # noqa: E501
            self._limpiar_todo()                                              # Limpia UI                                                          # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        self.curso_sel_id = int(info["curso_id"])                             # Set curso_id                                                       # noqa: E501
        self.inscripcion_sel_id = None                                        # Resetea selección                                                  # noqa: E501
        self._refrescar_inscritos()                                           # Carga inscritos                                                    # noqa: E501
        self._actualizar_suma_porcentajes()                                   # Suma %                                                             # noqa: E501
        self._limpiar_notas()                                                 # Limpia notas hasta elegir alumno                                   # noqa: E501
        self._status(f"Vista: Notas | CursoID: {self.curso_sel_id}")          # Estado                                                             # noqa: E501

    # =========================================================
    # Inscritos
    # =========================================================
    def _refrescar_inscritos(self):                                           # Recarga lista de inscritos                                         # noqa: E501
        for i in self.tree_insc.get_children():                               # Limpia tabla                                                       # noqa: E501
            self.tree_insc.delete(i)                                          # Borra                                                              # noqa: E501

        if self.curso_sel_id is None:                                         # Sin curso                                                          # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        rows = self._q(                                                       # Query inscritos + promedio                                         # noqa: E501
            """
            SELECT
              i.inscripcion_id,
              a.rut,
              a.nombres,
              a.apellidos,
              COALESCE(v.promedio_ponderado, 0) AS promedio
            FROM inscripciones i
            JOIN alumnos a ON a.alumno_id=i.alumno_id
            LEFT JOIN vw_promedios_ponderados v ON v.inscripcion_id=i.inscripcion_id
            WHERE i.curso_id=?
            ORDER BY a.apellidos, a.nombres
            """,
            (self.curso_sel_id,),
            fetch=True,
        ) or []

        for r in rows:                                                        # Inserta filas                                                      # noqa: E501
            alumno = f"{r['apellidos']} {r['nombres']}".strip()               # Nombre completo                                                    # noqa: E501
            self.tree_insc.insert(                                            # Inserta                                                            # noqa: E501
                "",
                "end",
                values=(int(r["inscripcion_id"]), r["rut"], alumno, f"{float(r['promedio']):.2f}"),
            )

    def on_select_inscrito(self):                                             # Selección de alumno inscrito                                       # noqa: E501
        sel = self.tree_insc.selection()                                      # Selección                                                          # noqa: E501
        if not sel:                                                           # Si no hay                                                          # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        vals = self.tree_insc.item(sel[0], "values")                          # Valores                                                            # noqa: E501
        if not vals:                                                          # Si vacío                                                           # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        self.inscripcion_sel_id = int(vals[0])                                # Set inscripción                                                    # noqa: E501
        self.lbl_alumno.config(text=f"{vals[2]} (RUT: {vals[1]})")            # Header alumno                                                      # noqa: E501

        self._cargar_notas_en_tabla()                                         # Carga evaluaciones + notas                                         # noqa: E501
        self.on_recalcular()                                                  # Actualiza promedio                                                 # noqa: E501

    # =========================================================
    # Tabla notas (evaluaciones + nota)
    # =========================================================
    def _limpiar_notas(self):                                                 # Limpia tabla notas                                                 # noqa: E501
        for i in self.tree_notas.get_children():                              # Recorre items                                                      # noqa: E501
            self.tree_notas.delete(i)                                         # Borra                                                              # noqa: E501
        self.lbl_alumno.config(text="Seleccione un alumno inscrito.")         # Texto                                                              # noqa: E501
        self.lbl_prom.config(text="Promedio ponderado: -")                    # Texto                                                              # noqa: E501
        self.on_limpiar_panel()                                               # Limpia panel                                                       # noqa: E501

    def _cargar_notas_en_tabla(self):                                         # Carga evaluaciones del curso + notas del alumno                    # noqa: E501
        self._limpiar_notas()                                                 # Limpia antes                                                       # noqa: E501
        if self.curso_sel_id is None or self.inscripcion_sel_id is None:      # Debe haber curso y alumno                                          # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        rows = self._q(                                                       # ✅ Query correcta para ver notas guardadas                          # noqa: E501
            """
            SELECT
              e.evaluacion_id,
              e.nombre,
              e.porcentaje,
              n.nota
            FROM evaluaciones e
            LEFT JOIN notas n
              ON n.evaluacion_id=e.evaluacion_id
             AND n.inscripcion_id=?
            WHERE e.curso_id=?
            ORDER BY e.evaluacion_id
            """,
            (self.inscripcion_sel_id, self.curso_sel_id),
            fetch=True,
        ) or []

        for r in rows:                                                        # Inserta filas                                                      # noqa: E501
            nota = r["nota"]                                                  # Puede ser None                                                     # noqa: E501
            nota_txt = "" if nota is None else f"{float(nota):.2f}"           # Muestra vacío si no existe                                         # noqa: E501
            self.tree_notas.insert(                                           # Inserta                                                            # noqa: E501
                "",
                "end",
                values=(int(r["evaluacion_id"]), r["nombre"], f"{float(r['porcentaje']):.2f}", nota_txt),
            )

    # =========================================================
    # Panel casillas
    # =========================================================
    def on_select_evaluacion(self):                                           # Al seleccionar evaluación en la tabla                              # noqa: E501
        sel = self.tree_notas.selection()                                     # Selección                                                          # noqa: E501
        if not sel:                                                           # Si no hay                                                          # noqa: E501
            return                                                            # Sale                                                               # noqa: E501
        vals = self.tree_notas.item(sel[0], "values")                         # Valores                                                            # noqa: E501
        if not vals:                                                          # Si vacío                                                           # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        self.var_eval_nombre.set(vals[1])                                     # Nombre                                                            # noqa: E501
        self.var_eval_porcentaje.set(vals[2])                                 # Porcentaje                                                         # noqa: E501
        self.var_nota.set(vals[3])                                            # Nota (puede estar vacía)                                           # noqa: E501

    def on_guardar_nota_panel(self):                                          # Guarda nota desde panel                                            # noqa: E501
        if self.inscripcion_sel_id is None:                                   # Debe seleccionar alumno                                            # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un alumno inscrito.")  # Aviso                                                         # noqa: E501
            return                                                            # Sale                                                               # noqa: E501
        sel = self.tree_notas.selection()                                     # Debe seleccionar evaluación                                        # noqa: E501
        if not sel:                                                           # Si no hay                                                          # noqa: E501
            messagebox.showwarning("Atención", "Seleccione una evaluación.")  # Aviso                                                              # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        item = sel[0]                                                         # Item                                                                # noqa: E501
        eval_id = int(self.tree_notas.set(item, "eval_id"))                   # EvalID                                                             # noqa: E501
        nota_str = (self.var_nota.get().strip() or "")                        # Nota string                                                        # noqa: E501

        try:
            if nota_str == "":                                                # Si vacío -> borrar nota                                            # noqa: E501
                self._q("DELETE FROM notas WHERE inscripcion_id=? AND evaluacion_id=?", (self.inscripcion_sel_id, eval_id))  # Delete # noqa: E501
                self.tree_notas.set(item, "nota", "")                         # Limpia celda                                                       # noqa: E501
            else:
                nota = float(nota_str.replace(",", "."))                      # Convierte                                                          # noqa: E501
                if nota < 0 or nota > 7:                                      # Rango chileno típico                                               # noqa: E501
                    raise ValueError("La nota debe estar entre 0 y 7.")       # Error                                                              # noqa: E501

                self._q(                                                      # UPSERT nota (si existe actualiza)                                  # noqa: E501
                    """
                    INSERT INTO notas(inscripcion_id, evaluacion_id, nota)
                    VALUES(?,?,?)
                    ON CONFLICT(inscripcion_id, evaluacion_id)
                    DO UPDATE SET nota=excluded.nota
                    """,
                    (self.inscripcion_sel_id, eval_id, nota),
                )
                self.tree_notas.set(item, "nota", f"{nota:.2f}")              # Actualiza tabla                                                    # noqa: E501

            self._refrescar_inscritos()                                       # Refresca promedios en lista izq                                    # noqa: E501
            self.on_recalcular()                                              # Refresca promedio                                                  # noqa: E501
            messagebox.showinfo("OK", "Nota guardada.")                       # OK                                                                 # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                             # Error                                                              # noqa: E501

    def on_limpiar_panel(self):                                               # Limpia panel                                                       # noqa: E501
        self.var_eval_nombre.set("")                                          # Limpia                                                            # noqa: E501
        self.var_eval_porcentaje.set("")                                      # Limpia                                                            # noqa: E501
        self.var_nota.set("")                                                 # Limpia                                                            # noqa: E501

    # =========================================================
    # Doble click: editor in-place en columna Nota
    # =========================================================
    def on_doble_click_nota(self, event):                                     # Doble click en tabla                                               # noqa: E501
        if self.inscripcion_sel_id is None:                                   # Debe haber alumno                                                  # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        row_id = self.tree_notas.identify_row(event.y)                        # Identifica fila                                                   # noqa: E501
        col = self.tree_notas.identify_column(event.x)                        # Identifica columna                                                # noqa: E501

        if col != "#4" or not row_id:                                         # Solo columna Nota (#4)                                            # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        bbox = self.tree_notas.bbox(row_id, col)                              # Caja de celda                                                     # noqa: E501
        if not bbox:                                                          # Si no hay                                                         # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        x, y, w, h = bbox                                                     # Coordenadas                                                       # noqa: E501

        if self._editor_entry is not None:                                    # Si ya existe un editor                                            # noqa: E501
            self._editor_entry.destroy()                                      # Lo destruye                                                       # noqa: E501

        valor_actual = self.tree_notas.set(row_id, "nota")                    # Valor actual                                                      # noqa: E501

        self._editor_entry = ttk.Entry(self.tree_notas)                       # Entry flotante                                                    # noqa: E501
        self._editor_entry.place(x=x, y=y, width=w, height=h)                 # Posiciona sobre celda                                             # noqa: E501
        self._editor_entry.insert(0, valor_actual)                            # Carga valor                                                       # noqa: E501
        self._editor_entry.focus()                                            # Foco                                                              # noqa: E501

        self._editor_item = row_id                                            # Guarda item editado                                               # noqa: E501

        self._editor_entry.bind("<Return>", lambda e: self._commit_edicion()) # Enter guarda                                                     # noqa: E501
        self._editor_entry.bind("<FocusOut>", lambda e: self._commit_edicion())  # Perder foco guarda                                             # noqa: E501

    def _commit_edicion(self):                                                # Confirma edición in-place                                          # noqa: E501
        if self._editor_entry is None or self._editor_item is None:           # Si no hay editor                                                  # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        nuevo = (self._editor_entry.get().strip() or "")                      # Nuevo valor                                                       # noqa: E501
        item = self._editor_item                                              # Item                                                              # noqa: E501

        try:
            eval_id = int(self.tree_notas.set(item, "eval_id"))               # EvalID                                                            # noqa: E501

            if nuevo == "":                                                   # Vacío => borrar nota                                              # noqa: E501
                self._q("DELETE FROM notas WHERE inscripcion_id=? AND evaluacion_id=?", (self.inscripcion_sel_id, eval_id))  # Delete # noqa: E501
                self.tree_notas.set(item, "nota", "")                         # Limpia celda                                                      # noqa: E501
                self.var_nota.set("")                                         # Panel también                                                     # noqa: E501
            else:
                nota = float(nuevo.replace(",", "."))                         # Convierte                                                         # noqa: E501
                if nota < 0 or nota > 7:                                      # Validación                                                        # noqa: E501
                    raise ValueError("La nota debe estar entre 0 y 7.")       # Error                                                              # noqa: E501

                self._q(                                                      # UPSERT                                                            # noqa: E501
                    """
                    INSERT INTO notas(inscripcion_id, evaluacion_id, nota)
                    VALUES(?,?,?)
                    ON CONFLICT(inscripcion_id, evaluacion_id)
                    DO UPDATE SET nota=excluded.nota
                    """,
                    (self.inscripcion_sel_id, eval_id, nota),
                )
                self.tree_notas.set(item, "nota", f"{nota:.2f}")              # Actualiza tabla                                                   # noqa: E501
                self.var_nota.set(f"{nota:.2f}")                              # Actualiza panel                                                   # noqa: E501

            self._refrescar_inscritos()                                       # Refresca promedios                                                # noqa: E501
            self.on_recalcular()                                              # Recalcula                                                         # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                             # Error                                                              # noqa: E501

        self._editor_entry.destroy()                                          # Cierra editor                                                     # noqa: E501
        self._editor_entry = None                                             # Limpia                                                            # noqa: E501
        self._editor_item = None                                              # Limpia                                                            # noqa: E501

    # =========================================================
    # Guardar tabla completa (recorre filas)
    # =========================================================
    def on_guardar_todo(self):                                                # Guarda todas las notas leyendo la tabla                            # noqa: E501
        if self.inscripcion_sel_id is None:                                   # Debe seleccionar alumno                                            # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un alumno inscrito.")  # Aviso                                                         # noqa: E501
            return                                                            # Sale                                                               # noqa: E501
        try:
            for item in self.tree_notas.get_children():                       # Recorre filas                                                      # noqa: E501
                eval_id = int(self.tree_notas.set(item, "eval_id"))           # EvalID                                                             # noqa: E501
                nota_txt = (self.tree_notas.set(item, "nota").strip() or "")  # Nota                                                                # noqa: E501

                if nota_txt == "":                                            # Si vacío                                                           # noqa: E501
                    self._q("DELETE FROM notas WHERE inscripcion_id=? AND evaluacion_id=?", (self.inscripcion_sel_id, eval_id))  # Delete # noqa: E501
                else:
                    nota = float(nota_txt.replace(",", "."))                  # Convierte                                                           # noqa: E501
                    if nota < 0 or nota > 7:                                  # Validación                                                         # noqa: E501
                        raise ValueError("La nota debe estar entre 0 y 7.")   # Error                                                              # noqa: E501
                    self._q(                                                  # UPSERT                                                            # noqa: E501
                        """
                        INSERT INTO notas(inscripcion_id, evaluacion_id, nota)
                        VALUES(?,?,?)
                        ON CONFLICT(inscripcion_id, evaluacion_id)
                        DO UPDATE SET nota=excluded.nota
                        """,
                        (self.inscripcion_sel_id, eval_id, nota),
                    )

            messagebox.showinfo("OK", "Notas guardadas (tabla completa).")    # OK                                                                 # noqa: E501
            self._refrescar_inscritos()                                       # Refresca promedios                                                # noqa: E501
            self.on_recalcular()                                              # Recalcula                                                         # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                             # Error                                                              # noqa: E501

    # =========================================================
    # Promedio ponderado
    # =========================================================
    def _actualizar_suma_porcentajes(self):                                   # Muestra suma % del curso                                           # noqa: E501
        if self.curso_sel_id is None:                                         # Sin curso                                                          # noqa: E501
            self.lbl_suma.config(text="Suma %: 0.00", foreground="black")     # Texto                                                              # noqa: E501
            return                                                            # Sale                                                               # noqa: E501
        row = self._q("SELECT COALESCE(SUM(porcentaje),0) AS s FROM evaluaciones WHERE curso_id=?", (self.curso_sel_id,), fetch=True)  # SUM # noqa: E501
        s = float(row[0]["s"] if row else 0)                                  # Suma                                                               # noqa: E501
        ok = abs(s - 100.0) < 0.001                                           # Valida 100                                                         # noqa: E501
        self.lbl_suma.config(text=f"Suma %: {s:.2f}", foreground=("green" if ok else "red"))  # Color                                        # noqa: E501

    def on_recalcular(self):                                                  # Recalcula promedio del alumno seleccionado                         # noqa: E501
        if self.inscripcion_sel_id is None:                                   # Sin alumno                                                         # noqa: E501
            self.lbl_prom.config(text="Promedio ponderado: -")                # Texto                                                              # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        row = self._q(                                                        # Consulta promedio de la vista                                      # noqa: E501
            "SELECT promedio_ponderado, suma_porcentajes FROM vw_promedios_ponderados WHERE inscripcion_id=?",
            (self.inscripcion_sel_id,),
            fetch=True,
        )
        if not row:                                                           # Si no hay vista/registro                                           # noqa: E501
            self.lbl_prom.config(text="Promedio ponderado: -")                # Texto                                                              # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        prom = float(row[0]["promedio_ponderado"] or 0)                        # Promedio                                                           # noqa: E501
        suma = float(row[0]["suma_porcentajes"] or 0)                          # Suma                                                               # noqa: E501

        if abs(suma - 100.0) > 0.001:                                         # Si suma distinta a 100                                             # noqa: E501
            self.lbl_prom.config(text=f"Promedio (⚠ suma %={suma:.2f}): {prom:.2f}")  # Texto                                                           # noqa: E501
        else:
            self.lbl_prom.config(text=f"Promedio ponderado: {prom:.2f}")      # Texto                                                              # noqa: E501

    # =========================================================
    # Inscribir / desinscribir
    # =========================================================
    def on_inscribir(self):                                                   # Inscribir alumno al curso                                          # noqa: E501
        if self.curso_sel_id is None:                                         # Debe seleccionar curso                                             # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un curso.")        # Aviso                                                              # noqa: E501
            return                                                            # Sale                                                               # noqa: E501

        win = tk.Toplevel(self)                                               # Ventana modal                                                     # noqa: E501
        win.title("Inscribir alumno")                                         # Título                                                            # noqa: E501
        win.geometry("720x480")                                               # Tamaño                                                            # noqa: E501
        win.transient(self)                                                   # Asociada                                                          # noqa: E501
        win.grab_set()                                                        # Modal                                                             # noqa: E501

        ttk.Label(win, text="Seleccione un alumno y presione 'Inscribir':", font=("Segoe UI", 10, "bold")).pack(pady=10)  # Texto # noqa: E501

        # Lista de alumnos
        rows = self._q(
            """
            SELECT alumno_id, rut, nombres, apellidos, COALESCE(email,'') AS email
            FROM alumnos
            WHERE estado=1
            ORDER BY apellidos, nombres
            """,
            fetch=True,
        ) or []

        tree = ttk.Treeview(win, columns=("id", "rut", "alumno", "email"), show="headings", selectmode="browse")  # Tree # noqa: E501
        tree.heading("id", text="ID")                                        # Head                                                             # noqa: E501
        tree.heading("rut", text="RUT")                                      # Head                                                             # noqa: E501
        tree.heading("alumno", text="Alumno")                                # Head                                                             # noqa: E501
        tree.heading("email", text="Email")                                  # Head                                                             # noqa: E501

        tree.column("id", width=60, anchor="center")                         # Col                                                              # noqa: E501
        tree.column("rut", width=120, anchor="w")                            # Col                                                              # noqa: E501
        tree.column("alumno", width=260, anchor="w")                         # Col                                                              # noqa: E501
        tree.column("email", width=200, anchor="w")                          # Col                                                              # noqa: E501

        for r in rows:                                                       # Inserta alumnos                                                   # noqa: E501
            nombre = f"{r['apellidos']} {r['nombres']}".strip()              # Nombre                                                            # noqa: E501
            tree.insert("", "end", values=(int(r["alumno_id"]), r["rut"], nombre, r["email"]))  # Insert                                   # noqa: E501

        tree.pack(fill="both", expand=True, padx=10, pady=10)                # Ubicación                                                         # noqa: E501

        def inscribir_sel():                                                 # Acción inscribir                                                  # noqa: E501
            sel = tree.selection()                                           # Selección                                                         # noqa: E501
            if not sel:                                                      # Si no hay                                                         # noqa: E501
                messagebox.showwarning("Atención", "Seleccione un alumno.")  # Aviso                                                             # noqa: E501
                return                                                       # Sale                                                              # noqa: E501
            vals = tree.item(sel[0], "values")                               # Values                                                            # noqa: E501
            alumno_id = int(vals[0])                                         # alumno_id                                                         # noqa: E501

            # Evita duplicado
            ex = self._q("SELECT 1 FROM inscripciones WHERE alumno_id=? AND curso_id=?", (alumno_id, self.curso_sel_id), fetch=True)  # Query # noqa: E501
            if ex:                                                           # Si existe                                                         # noqa: E501
                messagebox.showwarning("Atención", "El alumno ya está inscrito en este curso.")  # Aviso                               # noqa: E501
                return                                                       # Sale                                                              # noqa: E501

            try:
                self._q("INSERT INTO inscripciones(alumno_id, curso_id) VALUES(?,?)", (alumno_id, self.curso_sel_id))  # Insert     # noqa: E501
                messagebox.showinfo("OK", "Alumno inscrito.")                # OK                                                                 # noqa: E501
                win.destroy()                                                # Cierra                                                            # noqa: E501
                self._refrescar_inscritos()                                  # Refresca                                                          # noqa: E501
            except Exception as e:
                messagebox.showerror("Error", str(e))                        # Error                                                              # noqa: E501

        ttk.Button(win, text="Inscribir", command=inscribir_sel).pack(pady=10)  # Botón                                                          # noqa: E501

    def on_desinscribir(self):                                               # Quita inscripción                                                  # noqa: E501
        if self.inscripcion_sel_id is None:                                  # Debe seleccionar                                                   # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un alumno inscrito.")  # Aviso                                                         # noqa: E501
            return                                                            # Sale                                                               # noqa: E501
        if not messagebox.askyesno("Confirmar", "¿Quitar inscripción? (borra notas por cascada)"):  # Confirm                       # noqa: E501
            return                                                            # Cancel                                                             # noqa: E501
        try:
            self._q("DELETE FROM inscripciones WHERE inscripcion_id=?", (self.inscripcion_sel_id,))  # Delete                     # noqa: E501
            self.inscripcion_sel_id = None                                    # Limpia selección                                                   # noqa: E501
            self._refrescar_inscritos()                                       # Refresca                                                           # noqa: E501
            self._limpiar_notas()                                             # Limpia notas                                                        # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                             # Error                                                              # noqa: E501

    # =========================================================
    # Exportar (Excel/PDF): todas las notas del curso
    # =========================================================
    def _reporte_curso(self):                                                 # Prepara estructura reporte                                         # noqa: E501
        if self.curso_sel_id is None:                                         # Debe haber curso                                                   # noqa: E501
            raise ValueError("Seleccione un curso.")                          # Error                                                              # noqa: E501

        info = self._curso_info_sel()                                         # Dict curso                                                         # noqa: E501
        if not info:                                                          # Si no hay                                                          # noqa: E501
            raise ValueError("No se pudo obtener información del curso.")     # Error                                                              # noqa: E501

        evals = self._q("SELECT evaluacion_id, nombre, porcentaje FROM evaluaciones WHERE curso_id=? ORDER BY evaluacion_id", (self.curso_sel_id,), fetch=True) or []  # Evaluaciones # noqa: E501
        inscritos = self._q(                                                  # Inscritos                                                          # noqa: E501
            """
            SELECT i.inscripcion_id, a.rut, a.nombres, a.apellidos
            FROM inscripciones i
            JOIN alumnos a ON a.alumno_id=i.alumno_id
            WHERE i.curso_id=?
            ORDER BY a.apellidos, a.nombres
            """,
            (self.curso_sel_id,),
            fetch=True,
        ) or []

        # Trae todas las notas de ese curso en una consulta
        notas = self._q(                                                      # Notas por inscripción/evaluación                                  # noqa: E501
            """
            SELECT n.inscripcion_id, n.evaluacion_id, n.nota
            FROM notas n
            JOIN evaluaciones e ON e.evaluacion_id=n.evaluacion_id
            WHERE e.curso_id=?
            """,
            (self.curso_sel_id,),
            fetch=True,
        ) or []

        notas_map = {}                                                        # (insc_id, eval_id) -> nota                                        # noqa: E501
        for n in notas:                                                       # Recorre                                                            # noqa: E501
            notas_map[(int(n["inscripcion_id"]), int(n["evaluacion_id"]))] = float(n["nota"])  # Guarda                                             # noqa: E501

        # Suma porcentajes
        srow = self._q("SELECT COALESCE(SUM(porcentaje),0) AS s FROM evaluaciones WHERE curso_id=?", (self.curso_sel_id,), fetch=True)  # SUM # noqa: E501
        suma = float(srow[0]["s"] if srow else 0)                              # Suma                                                               # noqa: E501

        return info, evals, inscritos, notas_map, suma                         # Retorna todo                                                       # noqa: E501

    def on_exportar_excel(self):                                               # Exporta Excel completo                                             # noqa: E501
        try:
            info, evals, inscritos, notas_map, suma = self._reporte_curso()    # Datos                                                              # noqa: E501

            ruta = filedialog.asksaveasfilename(                               # Ruta                                                               # noqa: E501
                title="Guardar Excel (Notas del curso)",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                                       # Cancel                                                             # noqa: E501
                return                                                         # Sale                                                               # noqa: E501

            wb = Workbook()                                                    # Workbook                                                           # noqa: E501
            ws = wb.active                                                     # Hoja                                                               # noqa: E501
            ws.title = "Notas"                                                 # Nombre                                                             # noqa: E501

            # Encabezado informativo
            ws.append(["Universidad", info["universidad_nombre"]])             # Uni                                                                # noqa: E501
            ws.append(["Carrera", info["carrera_nombre"]])                     # Carrera                                                            # noqa: E501
            ws.append(["Periodo", info["periodo"]])                            # Periodo                                                            # noqa: E501
            ws.append(["Curso", info["curso_nombre"]])                         # Curso                                                              # noqa: E501
            ws.append(["Codigo", info.get("codigo", "")])                      # Código                                                             # noqa: E501
            ws.append(["Suma porcentajes", f"{suma:.2f}"])                     # Suma                                                               # noqa: E501
            ws.append([])                                                      # Línea en blanco                                                    # noqa: E501

            # Cabecera matriz
            headers = ["InscID", "RUT", "Apellidos", "Nombres"]                # Columnas fijas                                                     # noqa: E501
            for e in evals:                                                    # Columnas por evaluación                                            # noqa: E501
                headers.append(e["nombre"])                                    # Nombre eval                                                        # noqa: E501
            headers += ["PromedioPond"]                                        # Col promedio                                                       # noqa: E501
            ws.append(headers)                                                 # Cabecera                                                           # noqa: E501

            # Construye filas
            for ins in inscritos:                                              # Por cada alumno inscrito                                           # noqa: E501
                insc_id = int(ins["inscripcion_id"])                           # insc_id                                                            # noqa: E501
                row = [insc_id, ins["rut"], ins["apellidos"], ins["nombres"]]  # Parte fija                                                         # noqa: E501

                # Calcula promedio ponderado con lo que hay
                prom = 0.0                                                     # Acumulador                                                         # noqa: E501
                for e in evals:                                                # Recorre evaluaciones                                               # noqa: E501
                    eval_id = int(e["evaluacion_id"])                          # eval_id                                                            # noqa: E501
                    por = float(e["porcentaje"])                               # %                                                                   # noqa: E501
                    nota = notas_map.get((insc_id, eval_id), None)             # Nota o None                                                        # noqa: E501
                    row.append("" if nota is None else float(nota))            # En Excel: vacío si no hay                                          # noqa: E501
                    prom += (0.0 if nota is None else float(nota)) * por       # Suma ponderada                                                     # noqa: E501
                prom = prom / 100.0                                            # Divide por 100                                                     # noqa: E501
                row.append(round(prom, 2))                                     # Agrega promedio                                                     # noqa: E501
                ws.append(row)                                                 # Inserta fila                                                       # noqa: E501

            wb.save(ruta)                                                      # Guarda                                                             # noqa: E501
            messagebox.showinfo("OK", "Excel exportado con todas las notas.")  # OK                                                                 # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                              # Error                                                              # noqa: E501

    def on_exportar_pdf(self):                                                 # Exporta PDF (resumen tabla)                                        # noqa: E501
        try:
            info, evals, inscritos, notas_map, suma = self._reporte_curso()    # Datos                                                              # noqa: E501

            ruta = filedialog.asksaveasfilename(                               # Ruta                                                               # noqa: E501
                title="Guardar PDF (Notas del curso)",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialdir=carpeta_exports(),
            )
            if not ruta:                                                       # Cancel                                                             # noqa: E501
                return                                                         # Sale                                                               # noqa: E501

            c = canvas.Canvas(ruta, pagesize=A4)                               # Canvas                                                             # noqa: E501
            w, h = A4                                                          # Dimensiones                                                        # noqa: E501
            y = h - 60                                                         # Cursor Y                                                           # noqa: E501

            c.setFont("Helvetica-Bold", 13)                                    # Título                                                             # noqa: E501
            c.drawString(40, y, "Reporte de Notas (Curso)")                    # Texto                                                              # noqa: E501
            y -= 18                                                            # Baja                                                               # noqa: E501

            c.setFont("Helvetica", 10)                                         # Info                                                              # noqa: E501
            c.drawString(40, y, f"{info['universidad_nombre']} | {info['carrera_nombre']} | {info['periodo']} | {info['curso_nombre']}")  # Info # noqa: E501
            y -= 14                                                            # Baja                                                               # noqa: E501
            c.drawString(40, y, f"Suma porcentajes: {suma:.2f}%")              # Suma                                                              # noqa: E501
            y -= 20                                                            # Baja                                                               # noqa: E501

            # Cabecera simple (para no hacerlo enorme en PDF)
            c.setFont("Helvetica-Bold", 9)                                     # Fuente                                                            # noqa: E501
            c.drawString(40, y, "RUT")                                         # Col                                                                # noqa: E501
            c.drawString(120, y, "Alumno")                                     # Col                                                                # noqa: E501
            c.drawRightString(560, y, "Prom.")                                 # Col                                                                # noqa: E501
            y -= 12                                                            # Baja                                                               # noqa: E501

            c.setFont("Helvetica", 9)                                          # Fuente                                                            # noqa: E501
            for ins in inscritos:                                              # Recorre inscritos                                                 # noqa: E501
                if y < 60:                                                     # Salto página                                                      # noqa: E501
                    c.showPage()                                               # Nueva                                                             # noqa: E501
                    y = h - 60                                                 # Reinicia                                                          # noqa: E501
                    c.setFont("Helvetica", 9)                                  # Fuente                                                            # noqa: E501

                insc_id = int(ins["inscripcion_id"])                           # insc_id                                                           # noqa: E501
                alumno = f"{ins['apellidos']} {ins['nombres']}".strip()        # Nombre completo                                                   # noqa: E501

                prom = 0.0                                                     # Promedio ponderado                                                # noqa: E501
                for e in evals:                                                # Recorre evaluaciones                                              # noqa: E501
                    eval_id = int(e["evaluacion_id"])                          # eval_id                                                           # noqa: E501
                    por = float(e["porcentaje"])                               # porcentaje                                                        # noqa: E501
                    nota = notas_map.get((insc_id, eval_id), None)             # nota                                                              # noqa: E501
                    prom += (0.0 if nota is None else float(nota)) * por       # pondera                                                           # noqa: E501
                prom = prom / 100.0                                            # divide                                                           # noqa: E501

                c.drawString(40, y, str(ins["rut"])[:12])                      # RUT                                                               # noqa: E501
                c.drawString(120, y, alumno[:36])                              # Alumno                                                            # noqa: E501
                c.drawRightString(560, y, f"{prom:.2f}")                       # Prom                                                              # noqa: E501
                y -= 12                                                        # Baja                                                              # noqa: E501

            c.save()                                                           # Guarda                                                            # noqa: E501
            messagebox.showinfo("OK", "PDF exportado con resumen de promedios.")  # OK                                                            # noqa: E501
        except Exception as e:
            messagebox.showerror("Error", str(e))                              # Error                                                              # noqa: E501

    # =========================================================
    # Importar Excel de notas (2 formatos):
    # 1) Formato LARGO:  RUT | Evaluacion | Nota
    # 2) Formato MATRIZ: RUT | Apellidos | Nombres | <columnas evaluaciones...>
    # =========================================================
    def on_importar_excel(self):                                               # Importa notas del curso                                            # noqa: E501
        if self.curso_sel_id is None:                                          # Debe seleccionar curso                                             # noqa: E501
            messagebox.showwarning("Atención", "Seleccione un curso antes de importar notas.")  # Aviso                     # noqa: E501
            return                                                             # Sale                                                               # noqa: E501

        ruta = filedialog.askopenfilename(                                     # Selecciona archivo                                                 # noqa: E501
            title="Importar Excel (Notas del curso)",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=".",
        )
        if not ruta:                                                           # Cancel                                                             # noqa: E501
            return                                                             # Sale                                                               # noqa: E501

        try:
            wb = load_workbook(ruta)                                           # Abre libro                                                         # noqa: E501
            ws = wb.active                                                     # Hoja activa                                                        # noqa: E501

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Cabecera                                  # noqa: E501
            headers = [str(h or "").strip() for h in header_row]                # Headers                                                           # noqa: E501
            headers_cf = [h.casefold() for h in headers]                        # Headers normalizados                                              # noqa: E501

            def hidx(*names: str) -> int:                                      # Busca índice columna por nombres                                  # noqa: E501
                for n in names:                                                # Recorre alias                                                     # noqa: E501
                    k = n.casefold()                                           # Normaliza                                                         # noqa: E501
                    if k in headers_cf:                                        # Si existe                                                         # noqa: E501
                        return headers_cf.index(k)                             # Retorna índice                                                    # noqa: E501
                return -1                                                      # No encontrado                                                     # noqa: E501

            i_rut = hidx("rut")                                                # Índice RUT                                                        # noqa: E501
            i_eval = hidx("evaluacion", "evaluación", "evaluacion ")           # Índice evaluacion (formato largo)                                 # noqa: E501
            i_nota = hidx("nota")                                              # Índice nota (formato largo)                                       # noqa: E501

            # Map evaluación nombre->id
            evals = self._q("SELECT evaluacion_id, nombre FROM evaluaciones WHERE curso_id=?", (self.curso_sel_id,), fetch=True) or []  # Eval map # noqa: E501
            eval_map = {str(e["nombre"]).strip().casefold(): int(e["evaluacion_id"]) for e in evals}  # Mapa                                                            # noqa: E501

            # Mapa rut_normalizado->inscripcion_id (solo alumnos inscritos)
            ins_rows = self._q(                                                # Query inscritos                                                   # noqa: E501
                """
                SELECT i.inscripcion_id, a.rut_normalizado, a.rut
                FROM inscripciones i
                JOIN alumnos a ON a.alumno_id=i.alumno_id
                WHERE i.curso_id=?
                """,
                (self.curso_sel_id,),
                fetch=True,
            ) or []
            rut_to_insc = {str(r["rut_normalizado"]): int(r["inscripcion_id"]) for r in ins_rows}  # Mapa                                                           # noqa: E501

            total = 0                                                         # Total filas                                                       # noqa: E501
            ok = 0                                                            # OK                                                                # noqa: E501
            errores = []                                                      # Errores                                                           # noqa: E501

            # Si existe Evaluacion+Nota => formato largo
            if i_rut != -1 and i_eval != -1 and i_nota != -1:                 # Detecta formato largo                                             # noqa: E501
                for fila_n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Datos                   # noqa: E501
                    total += 1                                                # Cuenta                                                            # noqa: E501
                    try:
                        rut_val = str(row[i_rut] or "").strip()               # RUT                                                               # noqa: E501
                        ev_val = str(row[i_eval] or "").strip()               # Evaluación                                                        # noqa: E501
                        nota_val = row[i_nota]                                # Nota (puede ser número o texto)                                   # noqa: E501

                        if not rut_val or not ev_val:                         # Validación                                                        # noqa: E501
                            raise ValueError("Faltan RUT o Evaluacion.")      # Error                                                             # noqa: E501

                        # Normaliza rut como (numero+dv) sin guion
                        rut_norm = self._rut_norm_basico(rut_val)             # Normaliza                                                        # noqa: E501
                        insc_id = rut_to_insc.get(rut_norm)                   # Busca inscripción                                                # noqa: E501
                        if not insc_id:                                       # Si no está inscrito                                              # noqa: E501
                            raise ValueError("El RUT no está inscrito en este curso.")  # Error                                                     # noqa: E501

                        eval_id = eval_map.get(ev_val.casefold())             # EvalID por nombre                                                # noqa: E501
                        if not eval_id:                                       # Si no existe evaluación                                          # noqa: E501
                            raise ValueError("La evaluación no existe en este curso.")  # Error                                                     # noqa: E501

                        if nota_val is None or str(nota_val).strip() == "":   # Vacío => borrar nota                                             # noqa: E501
                            self._q("DELETE FROM notas WHERE inscripcion_id=? AND evaluacion_id=?", (insc_id, eval_id))  # Delete           # noqa: E501
                        else:
                            nota = float(str(nota_val).replace(",", "."))     # Convierte                                                        # noqa: E501
                            if nota < 0 or nota > 7:                          # Validación                                                       # noqa: E501
                                raise ValueError("Nota fuera de rango (0-7).")  # Error                                                         # noqa: E501
                            self._q(                                          # UPSERT nota                                                      # noqa: E501
                                """
                                INSERT INTO notas(inscripcion_id, evaluacion_id, nota)
                                VALUES(?,?,?)
                                ON CONFLICT(inscripcion_id, evaluacion_id)
                                DO UPDATE SET nota=excluded.nota
                                """,
                                (insc_id, eval_id, nota),
                            )

                        ok += 1                                               # OK                                                                # noqa: E501
                    except Exception as e:
                        errores.append(f"Fila {fila_n}: {e}")                 # Error fila                                                        # noqa: E501

            # Si NO es largo, intentamos matriz: RUT + columnas evaluaciones
            elif i_rut != -1:                                                 # Detecta matriz                                                    # noqa: E501
                # Columnas evaluación = todas las columnas desde la 4 (o desde donde no sea rut/apellidos/nombres)
                fixed = {"rut", "apellidos", "nombres", "inscid", "promediopond"}  # Fijas                                                          # noqa: E501
                eval_cols = []                                                # Lista columnas evaluaciones                                       # noqa: E501
                for idx, h in enumerate(headers):                             # Recorre cabecera                                                  # noqa: E501
                    if h.strip().casefold() in fixed:                         # Si es fija                                                       # noqa: E501
                        continue                                              # Salta                                                            # noqa: E501
                    if idx == i_rut:                                          # Si es rut                                                        # noqa: E501
                        continue                                              # Salta                                                            # noqa: E501
                    # Consideramos que cualquier columna no fija es evaluación
                    eval_cols.append((idx, h.strip()))                        # Guarda (index, nombre col)                                       # noqa: E501

                if not eval_cols:                                             # Sin columnas evaluación                                           # noqa: E501
                    raise ValueError("Formato matriz inválido: no se detectan columnas de evaluaciones.")  # Error                 # noqa: E501

                for fila_n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):        # Datos                   # noqa: E501
                    total += 1                                                # Cuenta                                                            # noqa: E501
                    try:
                        rut_val = str(row[i_rut] or "").strip()               # RUT                                                               # noqa: E501
                        if not rut_val:                                       # Validación                                                        # noqa: E501
                            raise ValueError("RUT vacío.")                    # Error                                                             # noqa: E501
                        rut_norm = self._rut_norm_basico(rut_val)             # Normaliza                                                        # noqa: E501
                        insc_id = rut_to_insc.get(rut_norm)                   # Busca inscripción                                                # noqa: E501
                        if not insc_id:                                       # No inscrito                                                      # noqa: E501
                            raise ValueError("El RUT no está inscrito en este curso.")  # Error                                                     # noqa: E501

                        # Para cada columna evaluación, upsert
                        for idx, eval_name in eval_cols:                      # Recorre eval cols                                                # noqa: E501
                            eval_id = eval_map.get(eval_name.casefold())      # Busca evaluacion_id                                              # noqa: E501
                            if not eval_id:                                   # Si no existe en BD                                               # noqa: E501
                                continue                                      # No rompe, ignora esa columna                                     # noqa: E501

                            val = row[idx]                                    # Valor celda nota                                                 # noqa: E501
                            if val is None or str(val).strip() == "":         # Vacío => borrar                                                  # noqa: E501
                                self._q("DELETE FROM notas WHERE inscripcion_id=? AND evaluacion_id=?", (insc_id, eval_id))  # Delete           # noqa: E501
                            else:
                                nota = float(str(val).replace(",", "."))      # Convierte                                                        # noqa: E501
                                if nota < 0 or nota > 7:                      # Valida                                                           # noqa: E501
                                    raise ValueError(f"Nota fuera de rango en '{eval_name}'.")  # Error                                                   # noqa: E501
                                self._q(                                      # UPSERT                                                           # noqa: E501
                                    """
                                    INSERT INTO notas(inscripcion_id, evaluacion_id, nota)
                                    VALUES(?,?,?)
                                    ON CONFLICT(inscripcion_id, evaluacion_id)
                                    DO UPDATE SET nota=excluded.nota
                                    """,
                                    (insc_id, eval_id, nota),
                                )

                        ok += 1                                               # OK fila                                                          # noqa: E501
                    except Exception as e:
                        errores.append(f"Fila {fila_n}: {e}")                 # Error fila                                                        # noqa: E501

            else:
                raise ValueError("Excel no tiene columna RUT válida.")        # Error                                                             # noqa: E501

            # Refresca UI
            self._refrescar_inscritos()                                       # Refresca lista                                                    # noqa: E501
            if self.inscripcion_sel_id is not None:                           # Si hay alumno seleccionado                                        # noqa: E501
                self._cargar_notas_en_tabla()                                 # Refresca notas                                                    # noqa: E501
                self.on_recalcular()                                          # Refresca promedio                                                 # noqa: E501

            msg = f"Importación finalizada.\nProcesadas: {total}\nOK: {ok}\nErrores: {len(errores)}"  # Resumen               # noqa: E501
            if errores:                                                       # Si hay errores                                                     # noqa: E501
                msg += "\n\nPrimeros errores:\n" + "\n".join(errores[:10])    # Muestra 10                                                        # noqa: E501
            messagebox.showinfo("Resultado", msg)                              # Muestra resultado                                                  # noqa: E501

        except Exception as e:
            messagebox.showerror("Error", str(e))                              # Error general                                                       # noqa: E501

    def _rut_norm_basico(self, rut: str) -> str:                               # Normaliza rut a numero+dv sin guion (básico)                        # noqa: E501
        t = (rut or "").strip().upper()                                       # Limpia y mayúsculas                                                # noqa: E501
        t = t.replace(".", "").replace(" ", "")                               # Quita puntos/espacios                                              # noqa: E501
        if "-" not in t and len(t) >= 2:                                      # Si no trae guion                                                   # noqa: E501
            t = t[:-1] + "-" + t[-1]                                          # Inserta guion                                                      # noqa: E501
        parts = t.split("-")                                                  # Separa                                                             # noqa: E501
        if len(parts) != 2:                                                   # Debe haber 2 partes                                                # noqa: E501
            return ""                                                         # Devuelve vacío                                                     # noqa: E501
        num = parts[0]                                                        # Número                                                             # noqa: E501
        dv = parts[1]                                                         # DV                                                                 # noqa: E501
        if not num.isdigit():                                                 # Debe ser numérico                                                  # noqa: E501
            return ""                                                         # Vacío                                                              # noqa: E501
        return f"{int(num)}{dv}"                                              # Normalizado                                                        # noqa: E501

    # =========================================================
    # Limpieza general
    # =========================================================
    def _limpiar_todo(self):                                                  # Limpia todo el contenido                                           # noqa: E501
        for i in self.tree_insc.get_children():                               # Limpia inscritos                                                  # noqa: E501
            self.tree_insc.delete(i)                                          # Borra                                                             # noqa: E501
        self._limpiar_notas()                                                 # Limpia notas                                                      # noqa: E501
        self.lbl_suma.config(text="Suma %: 0.00", foreground="black")         # Suma 0                                                            # noqa: E501
        self._status("Vista: Notas | Sin cursos")                             # Estado                                                            # noqa: E501
